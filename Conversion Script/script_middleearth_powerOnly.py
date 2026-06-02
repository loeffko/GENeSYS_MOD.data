# -*- coding: utf-8 -*-
"""Generate a POWER-ONLY Middle-Earth parameter Excel:
RegularParameters_MiddleEarth_powerOnly.xlsx.

Mirrors the NA power-only approach (script_northamerica.py): the filtered model
keeps only Power technologies, the single 'Power' fuel, and the four power
storages. Heat/Mobility/H2 etc. are excluded. Run together with
script_middleearth_allfuels.py to produce the matching all-fuels Excel that the
Julia precompute reads when switch_power_only_mode = 1.

This script builds Set_filter_file_MiddleEarth_powerOnly.xlsx fresh from the
base ME filter, then invokes master_function with parameters_only output. The
existing Timeseries_MiddleEarth.xlsx is reusable as-is (regions/years are the
same; excluded fuel demand entries are dropped at dataload).
"""
import os
import shutil
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
BASE_FILTER = os.path.join(HERE, "Set_filter_file_MiddleEarth.xlsx")
PWR_FILTER  = os.path.join(HERE, "Set_filter_file_MiddleEarth_powerOnly.xlsx")
OUTPUT_DIR  = os.path.normpath(os.path.join(HERE, "..", "Output", "output_excel"))

POWER_TECHS = {
    # Thermal + CHP power generation
    "P_Biomass", "P_Biomass_CCS", "P_CSP",
    "P_Coal_Hardcoal", "P_Coal_Hardcoal_CCS",
    "P_Coal_Lignite",  "P_Coal_Lignite_CCS",
    "P_Gas_CCGT", "P_Gas_CCS", "P_Gas_Engines", "P_Gas_OCGT",
    "P_Geothermal", "P_H2_OCGT",
    "P_Hydro_Reservoir", "P_Hydro_RoR",
    "P_Nuclear", "P_Ocean", "P_Oil",
    "P_PV_Rooftop_Commercial", "P_PV_Rooftop_Residential",
    "P_PV_Utility_Avg", "P_PV_Utility_Inf", "P_PV_Utility_Opt",
    "P_PV_Utility_Tracking",
    "P_Wind_Offshore_Deep", "P_Wind_Offshore_Shallow",
    "P_Wind_Offshore_Transitional",
    "P_Wind_Onshore_Avg", "P_Wind_Onshore_Inf", "P_Wind_Onshore_Opt",
    "CHP_Biomass_Solid", "CHP_Biomass_Solid_CCS",
    "CHP_Coal_Hardcoal", "CHP_Coal_Hardcoal_CCS",
    "CHP_Coal_Lignite",  "CHP_Coal_Lignite_CCS",
    "CHP_Gas_CCGT_Biogas", "CHP_Gas_CCGT_Biogas_CCS",
    "CHP_Gas_CCGT_Natural", "CHP_Gas_CCGT_Natural_CCS",
    "CHP_Gas_CCGT_SynGas",
    "CHP_Hydrogen_FuelCell", "CHP_Oil",
    # Power storage discharge techs
    "D_Battery_Li-Ion", "D_Battery_Redox", "D_CAES", "D_PHS",
}
POWER_FUELS    = {"Power"}
POWER_STORAGES = {"S_Battery_Li-Ion", "S_Battery_Redox", "S_CAES", "S_PHS"}


def set_selection(xlsx, sheet, enabled_set):
    """Set column-B flag = 1 for column-A names in `enabled_set`, 0 otherwise."""
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    es = {str(e) for e in enabled_set}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is not None:
            ws.cell(row=r, column=2).value = 1 if str(name) in es else 0
    wb.save(xlsx)


def _read_ts_selection(xlsx):
    """Return [(name, flag), ...] from the Timeseries_selection sheet, or None."""
    wb = openpyxl.load_workbook(xlsx)
    if "Timeseries_selection" not in wb.sheetnames:
        return None
    ws = wb["Timeseries_selection"]
    out = []
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if n is not None:
            out.append((str(n), v))
    return out


def _write_ts_selection(xlsx, entries):
    """Overwrite the Timeseries_selection sheet with `entries` ([(name, flag),...])."""
    wb = openpyxl.load_workbook(xlsx)
    if "Timeseries_selection" in wb.sheetnames:
        del wb["Timeseries_selection"]
    ws = wb.create_sheet("Timeseries_selection")
    ws.cell(1, 1, "Timeseries")
    ws.cell(1, 2, "Timeseries selected")
    for i, (n, v) in enumerate(entries, start=2):
        ws.cell(i, 1, n)
        ws.cell(i, 2, v if v is not None else 1)
    wb.save(xlsx)


def main():
    # 1) Build the power-only ME filter from the existing ME filter (keep its
    #    Regions/Years/Modes/Emissions/etc; narrow Tech/Fuel/Storage)
    # Preserve any user-edited Timeseries_selection across the copy: the
    # power-only filter file is the source of truth for which TS sheets get
    # produced — only the tech/fuel/storage narrowing is rebuilt every run.
    saved_ts_sel = _read_ts_selection(PWR_FILTER) if os.path.exists(PWR_FILTER) else None
    shutil.copyfile(BASE_FILTER, PWR_FILTER)
    set_selection(PWR_FILTER, "Technology_selection", POWER_TECHS)
    set_selection(PWR_FILTER, "Fuel_selection",       POWER_FUELS)
    set_selection(PWR_FILTER, "Storage_selection",    POWER_STORAGES)
    if saved_ts_sel is not None:
        _write_ts_selection(PWR_FILTER, saved_ts_sel)
    print(f"Built {os.path.basename(PWR_FILTER)}: "
          f"{len(POWER_TECHS)} techs / {len(POWER_FUELS)} fuels / "
          f"{len(POWER_STORAGES)} storages")

    # 2) Run full conversion (params + timeseries). The power-only filter has a
    #    Timeseries_selection that drops Heat/Cool/HP/Mobility TS sheets so the
    #    generated Timeseries Excel stays lean.
    from functions.function_import import master_function
    master_function(os.path.basename(PWR_FILTER),
                    "excel", "long", "both",
                    "MiddleEarth", False, "Gondor")

    # 3) Rename outputs so power-only files sit next to the full ones
    for base in ("RegularParameters_MiddleEarth", "Timeseries_MiddleEarth"):
        src = os.path.join(OUTPUT_DIR, f"{base}.xlsx")
        dst = os.path.join(OUTPUT_DIR, f"{base}_powerOnly.xlsx")
        if not os.path.exists(src):
            continue
        if os.path.exists(dst):
            os.remove(dst)
        os.rename(src, dst)
        print(f"Generated: {dst}")


if __name__ == "__main__":
    main()
