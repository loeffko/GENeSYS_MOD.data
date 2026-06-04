# -*- coding: utf-8 -*-
"""Generate an UNFILTERED (all fuels + all techs + all storages) Middle-Earth
parameter Excel: RegularParameters_MiddleEarth_allFuels.xlsx.

Counterpart to script_middleearth_powerOnly.py: when switch_power_only_mode = 1
the Julia precompute reads both files. The all-fuels Excel contains the fossil
supply chain (Z_Import_*, fuel IAR, EmissionContentPerFuel) that the power-only
Excel filters out — the precompute uses it to bake an effective per-tech
VariableCost and OutputEmissionRatio into the power-only Parameters.

Builds Set_filter_file_MiddleEarth_allFuels.xlsx fresh from the universal
filter (toggled to ME regions/years), then enables every entry from the base
Sets_*.csv files (catches techs/fuels added after the universal filter was last
refreshed, e.g. the Z_Import_Biomass/Lignite dummies).
"""
import os
import shutil
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
BASE_FILTER = os.path.join(HERE, "Set_filter_file_MiddleEarth.xlsx")  # ME regions/years preset
ALL_FILTER  = os.path.join(HERE, "Set_filter_file_MiddleEarth_allFuels.xlsx")
OUTPUT_DIR  = os.path.normpath(os.path.join(HERE, "..", "Output", "output_excel"))
SETS_DIR    = os.path.normpath(os.path.join(HERE, "..", "Data", "Parameters", "00_Sets&Tags"))


def set_selection(xlsx, sheet, enabled_set):
    """Set column-B flag = 1 for column-A names in `enabled_set`, 0 otherwise."""
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    es = {str(e) for e in enabled_set}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is not None:
            ws.cell(row=r, column=2).value = 1 if str(name) in es else 0
    wb.save(xlsx)


def ensure_all_enabled(xlsx, sheet, names):
    """Set every existing column-A entry to 1; APPEND any name in `names` not yet
    present (also enabled). Catches techs/fuels added to base Sets CSVs after
    the universal filter file was last refreshed."""
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    present = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            present.add(str(v))
            ws.cell(row=r, column=2).value = 1
    next_row = ws.max_row + 1
    for n in names:
        if str(n) not in present:
            ws.cell(row=next_row, column=1).value = n
            ws.cell(row=next_row, column=2).value = 1
            next_row += 1
            present.add(str(n))
    wb.save(xlsx)


def read_set(filename):
    """Read a single-column Sets_*.csv (headerless or with header). Returns the
    list of non-empty stripped strings, skipping the obvious header row."""
    path = os.path.join(SETS_DIR, filename)
    out = []
    with open(path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            v = line.strip().split(",")[0].strip()
            if not v:
                continue
            if i == 0 and v in ("Technology", "Fuel", "Storage", "Region", "Year"):
                continue
            out.append(v)
    return out


def main():
    # Start from the ME filter (preserves correct Regions/Years/etc) and enable
    # every Technology/Fuel/Storage on top — including entries added to the base
    # Sets CSVs after the ME filter was last refreshed (e.g. Z_Import dummies).
    shutil.copyfile(BASE_FILTER, ALL_FILTER)
    ensure_all_enabled(ALL_FILTER, "Fuel_selection",       read_set("Sets_Fuel.csv"))
    ensure_all_enabled(ALL_FILTER, "Technology_selection", read_set("Sets_Technology.csv"))
    ensure_all_enabled(ALL_FILTER, "Storage_selection",    read_set("Sets_Storage.csv"))
    print(f"Built {os.path.basename(ALL_FILTER)} "
          f"(ME regions/years preserved; all fuels/techs/storages enabled)")

    # Preserve any existing full-ME params Excel so this run does not clobber
    # it (master_function writes to RegularParameters_MiddleEarth.xlsx based on
    # scenario_option).
    regular_path = os.path.join(OUTPUT_DIR, "RegularParameters_MiddleEarth.xlsx")
    regular_bak  = regular_path + ".full.bak"
    if os.path.exists(regular_path):
        shutil.move(regular_path, regular_bak)

    try:
        from functions.function_import import master_function
        master_function(os.path.basename(ALL_FILTER),
                        "excel", "long", "parameters_only",
                        "MiddleEarth", False, "Gondor")

        src = os.path.join(OUTPUT_DIR, "RegularParameters_MiddleEarth.xlsx")
        dst = os.path.join(OUTPUT_DIR, "RegularParameters_MiddleEarth_allFuels.xlsx")
        if os.path.exists(dst):
            os.remove(dst)
        os.rename(src, dst)
        print(f"Generated: {dst}")
    finally:
        if os.path.exists(regular_bak):
            shutil.move(regular_bak, regular_path)
            print(f"Restored full-ME Excel from {os.path.basename(regular_bak)}")


if __name__ == "__main__":
    main()
