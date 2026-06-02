# -*- coding: utf-8 -*-
"""Generate an UNFILTERED (all fuels + all techs + all storages) North-America
parameter Excel: RegularParameters_NorthAmerica_allFuels.xlsx.

Used by the power-only precompute path (path 2): the Julia model reads both the
power-only Excel and this all-fuels Excel; for each thermal power tech it pulls
the input-fuel IAR/OAR + the upstream Z_Import_*.VariableCost +
EmissionContentPerFuel from this file to derive an effective output-side
VariableCost and EmissionActivityRatio.

NA regions + years 2025-2040 selected; all fuels/techs/storages enabled.
Power-only NA filter is backed up + restored, so the power-only conversion
(script_northamerica.py) is unaffected.
"""
import os
import shutil
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
UNIVERSAL = os.path.join(HERE, "Set_filter_file.xlsx")
NA_FILTER = os.path.join(HERE, "Set_filter_file_NorthAmerica.xlsx")
ALL_FILTER = os.path.join(HERE, "Set_filter_file_NorthAmerica_allFuels.xlsx")
OUTPUT_DIR = os.path.normpath(os.path.join(HERE, "..", "Output", "output_excel"))
SETS_DIR = os.path.normpath(os.path.join(HERE, "..", "Data", "Parameters", "00_Sets&Tags"))

NA_REGIONS = ["California", "WECC", "SPP", "MISO", "ERCOT", "SERC", "PJM",
              "NewYork", "NewEngland", "Canada"]
NA_YEARS = [str(y) for y in range(2025, 2041)]


def set_selection(xlsx, sheet, enabled_set):
    """Set column-B flag = 1 for column-A names in `enabled_set`, 0 otherwise."""
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    es = {str(e) for e in enabled_set}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is not None:
            ws.cell(row=r, column=2).value = 1 if str(name) in es else 0
    wb.save(xlsx)


def ensure_all_enabled(xlsx, sheet, names):
    """Set every existing column-A entry to 1; APPEND any name in `names` not yet
    present (also enabled). Catches techs/fuels added to base Sets CSVs after the
    universal filter file was last refreshed."""
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    present = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            present.add(str(v))
            ws.cell(row=r, column=2).value = 1
    next_row = ws.max_row + 1
    for n in names:
        if str(n) not in present:
            ws.cell(row=next_row, column=1).value = n
            ws.cell(row=next_row, column=2).value = 1
            next_row += 1
            present.add(str(n))
    wb.save(xlsx)


def read_set(filename):
    """Read a single-column Sets_*.csv (headerless or with header). Returns list of
    non-empty stripped strings, skipping any obvious header row."""
    path = os.path.join(SETS_DIR, filename)
    out = []
    with open(path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            v = line.strip().split(",")[0].strip()
            if not v:
                continue
            # skip header row (Sets_Technology has 'Technology' header; Sets_Fuel is headerless)
            if i == 0 and v in ("Technology", "Fuel", "Storage", "Region", "Year"):
                continue
            out.append(v)
    return out


def main():
    # 1) Build the all-fuels NA filter from the universal one
    shutil.copyfile(UNIVERSAL, ALL_FILTER)
    set_selection(ALL_FILTER, "Region_selection", NA_REGIONS)
    set_selection(ALL_FILTER, "Year_selection", NA_YEARS)
    ensure_all_enabled(ALL_FILTER, "Fuel_selection",       read_set("Sets_Fuel.csv"))
    ensure_all_enabled(ALL_FILTER, "Technology_selection", read_set("Sets_Technology.csv"))
    ensure_all_enabled(ALL_FILTER, "Storage_selection",    read_set("Sets_Storage.csv"))
    print(f"Built {os.path.basename(ALL_FILTER)} (NA regions/years; all fuels/techs/storages enabled)")

    # 2) Swap the power-only NA filter out, install the all-fuels one (master_function
    #    reads 'Set_filter_file_NorthAmerica.xlsx' by name)
    backup = NA_FILTER + ".powerOnly.bak"
    if os.path.exists(NA_FILTER):
        shutil.copyfile(NA_FILTER, backup)
    shutil.copyfile(ALL_FILTER, NA_FILTER)

    try:
        # 3) Run conversion (params only, no timeseries needed)
        from functions.function_import import master_function
        master_function("Set_filter_file_NorthAmerica.xlsx", "excel", "long",
                        "parameters_only", "NorthAmerica", False, "California")

        # 4) Rename output to _allFuels (so the power-only Excel isn't clobbered)
        src = os.path.join(OUTPUT_DIR, "RegularParameters_NorthAmerica.xlsx")
        dst = os.path.join(OUTPUT_DIR, "RegularParameters_NorthAmerica_allFuels.xlsx")
        if os.path.exists(dst):
            os.remove(dst)
        os.rename(src, dst)
        print(f"Generated: {dst}")
    finally:
        # 5) Always restore the power-only NA filter
        if os.path.exists(backup):
            shutil.copyfile(backup, NA_FILTER)
            print(f"Restored power-only NA filter from {os.path.basename(backup)}")


if __name__ == "__main__":
    main()
