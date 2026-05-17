"""
expand_data.py
--------------
Reusable, config-driven helpers to expand the GENeSYS-MOD data set in the
``GENeSYS_MOD.data`` repository.

The functions are deliberately generic so the same code serves future needs
(adding technologies, fuels, storages, ...) and not only regions:

  * duplicate_member_in_df()   copy every row of one set member into a new one
  * duplicate_regions_in_df()  region-aware copy (handles trade region-pairs)
  * interpolate_years_in_df()  add linearly-interpolated years to a parameter
  * append_set_entries()       add names to a Sets_*.csv
  * add_timeseries_columns()   add region columns to wide TS_*.csv files
  * update_selection_sheet()   enable/disable/add rows in Set_filter_file.xlsx

Parameter CSVs are long-format:
    <index cols...>, Value, <blank>, Unit, Source, Updated at, Updated by
"Index columns" = every column left of ``Value``. Region-pair (trade)
parameters carry two region columns: ``Region`` and ``Region.1``.
"""

import os
import pandas as pd
import numpy as np
import openpyxl

DUMMY_SOURCE = "dummy data - empty entry"
INTERP_SOURCE = "interpolated"


# ----------------------------------------------------------------------
# Parameter CSV discovery / IO
# ----------------------------------------------------------------------
def param_csv_paths(parameters_dir):
    """Return the master ``Par_X/Par_X.csv`` files (scenario subfolders skipped)."""
    out = []
    for name in sorted(os.listdir(parameters_dir)):
        d = os.path.join(parameters_dir, name)
        csv = os.path.join(d, name + ".csv")
        if os.path.isdir(d) and os.path.isfile(csv):
            out.append(csv)
    return out


def read_param(csv_path):
    """Read a parameter CSV as text (keeps untouched rows byte-stable on write)."""
    return pd.read_csv(csv_path, dtype=str, keep_default_na=False, encoding="utf-8")


def _detect_newline(path):
    r"""Return the newline string an existing text file uses ('\n' default)."""
    try:
        with open(path, "rb") as f:
            chunk = f.read(65536)
        return "\r\n" if b"\r\n" in chunk else "\n"
    except FileNotFoundError:
        return "\n"


def _to_csv(df, file_obj, newline):
    """to_csv with an explicit line terminator (compatible across pandas versions)."""
    try:
        df.to_csv(file_obj, index=False, lineterminator=newline)
    except TypeError:                       # pandas < 1.5
        df.to_csv(file_obj, index=False, line_terminator=newline)


def write_param(df, csv_path):
    """Write a parameter CSV, preserving the file's original line endings."""
    newline = _detect_newline(csv_path)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        _to_csv(df, f, newline)


def index_columns(df):
    """Columns left of ``Value`` - the parameter's index/dimension columns."""
    if "Value" not in df.columns:
        return list(df.columns)
    return list(df.columns[:df.columns.get_loc("Value")])


def region_columns(df):
    """Region dimension columns ('Region', plus 'Region.1' for trade pairs)."""
    return [c for c in index_columns(df)
            if c == "Region" or c.startswith("Region.")]


# ----------------------------------------------------------------------
# Duplicating set members (regions, technologies, fuels, ...)
# ----------------------------------------------------------------------
def duplicate_member_in_df(df, column, source, new, source_tag=DUMMY_SOURCE):
    """
    Copy every row where ``df[column] == source`` into new rows with
    ``column == new``. Generic: use for technologies, fuels, storages, ...
    """
    if column not in df.columns:
        return df
    sub = df[df[column] == source].copy()
    if sub.empty:
        return df
    sub[column] = new
    if "Source" in sub.columns:
        sub["Source"] = source_tag
    return pd.concat([df, sub], ignore_index=True)


def duplicate_regions_in_df(df, region_map, source_tag=DUMMY_SOURCE):
    """
    Add new regions by copying their mapped source region's rows.

    region_map : {new_region: source_region}, assumed 1:1 (bijection).

    Single-region parameters  -> straight row copy with the region relabelled.
    Trade region-pair params  -> copy ordered pairs whose BOTH endpoints are
    mapped source regions, relabelling both endpoints. This mirrors the
    intra-source trade graph onto the new regions.
    """
    rcols = region_columns(df)
    if not rcols:
        return df
    inv = {src: new for new, src in region_map.items()}   # source -> new
    sources = set(region_map.values())
    add = []
    if len(rcols) == 1:
        col = rcols[0]
        for new_r, src_r in region_map.items():
            sub = df[df[col] == src_r].copy()
            if sub.empty:
                continue
            sub[col] = new_r
            if "Source" in sub.columns:
                sub["Source"] = source_tag
            add.append(sub)
    else:
        c1, c2 = rcols[0], rcols[1]
        for new_r, src_r in region_map.items():
            sub = df[(df[c1] == src_r) & (df[c2].isin(sources))].copy()
            if sub.empty:
                continue
            sub[c2] = sub[c2].map(inv)
            sub[c1] = new_r
            if "Source" in sub.columns:
                sub["Source"] = source_tag
            add.append(sub)
    if not add:
        return df
    return pd.concat([df] + add, ignore_index=True)


# ----------------------------------------------------------------------
# Year interpolation
# ----------------------------------------------------------------------
def interpolate_years_in_df(df, target_years, source_tag=INTERP_SOURCE):
    """
    For every index group, add a row for each year in ``target_years`` that is
    missing AND lies within the group's existing [min, max] year range, with
    ``Value`` linearly interpolated between the surrounding data points.

    Rows whose ``Year`` is the literal ``All`` are left untouched. Groups with
    fewer than two numeric data points, or any non-numeric value, are skipped.
    """
    if "Year" not in df.columns or "Value" not in df.columns:
        return df
    yr = df["Year"].astype(str).str.strip()
    df_yr = df[yr.str.lower() != "all"].copy()
    if df_yr.empty:
        return df

    df_yr["_Y"] = pd.to_numeric(df_yr["Year"], errors="coerce")
    df_yr["_V"] = pd.to_numeric(df_yr["Value"], errors="coerce")
    df_yr = df_yr[df_yr["_Y"].notna()]
    if df_yr.empty:
        return df

    key_cols = [c for c in index_columns(df) if c != "Year"]
    add = []
    for _, g in df_yr.groupby(key_cols, dropna=False, sort=False):
        g = g.sort_values("_Y")
        years = g["_Y"].to_numpy()
        vals = g["_V"].to_numpy()
        if len(years) < 2 or np.isnan(vals).any():
            continue
        have = set(int(y) for y in years)
        template = g.iloc[0].drop(["_Y", "_V"])
        for ty in target_years:
            if ty in have or ty < years[0] or ty > years[-1]:
                continue
            row = template.copy()
            row["Year"] = str(ty)
            row["Value"] = "%g" % float(np.interp(ty, years, vals))
            if "Source" in row.index:
                row["Source"] = source_tag
            add.append(row)
    if not add:
        return df
    return pd.concat([df, pd.DataFrame(add)], ignore_index=True)


# ----------------------------------------------------------------------
# Sets_*.csv (plain single-column lists)
# ----------------------------------------------------------------------
def append_set_entries(sets_csv_path, entries):
    """Append entries to a Sets_*.csv (single column), skipping duplicates."""
    df = pd.read_csv(sets_csv_path, dtype=str, keep_default_na=False, encoding="utf-8")
    col = df.columns[0]
    have = set(df[col].astype(str))
    new = [e for e in entries if str(e) not in have]
    if new:
        newline = _detect_newline(sets_csv_path)
        df = pd.concat([df, pd.DataFrame({col: [str(e) for e in new]})],
                       ignore_index=True)
        with open(sets_csv_path, "w", newline="", encoding="utf-8") as f:
            _to_csv(df, f, newline)
    return new


# ----------------------------------------------------------------------
# Timeseries TS_*.csv (wide: HOUR + one column per region)
# ----------------------------------------------------------------------
def timeseries_csv_paths(timeseries_dir):
    """Master TS_X/TS_X.csv files (subfolders such as MiddleEarth are skipped)."""
    out = []
    for name in sorted(os.listdir(timeseries_dir)):
        d = os.path.join(timeseries_dir, name)
        csv = os.path.join(d, name + ".csv")
        if os.path.isdir(d) and os.path.isfile(csv):
            out.append(csv)
    return out


def add_timeseries_columns(ts_csv_path, region_map):
    """
    Add a column per new region to a wide TS_*.csv, copied from its source
    region's column. Row 0 (the 'Source: ...' metadata line) is preserved.
    """
    newline = _detect_newline(ts_csv_path)
    with open(ts_csv_path, "r", newline="", encoding="utf-8") as f:
        meta_line = f.readline()            # keeps the original line ending
    df = pd.read_csv(ts_csv_path, skiprows=1, dtype=str, keep_default_na=False,
                     encoding="utf-8")
    added = []
    for new_r, src_r in region_map.items():
        if src_r in df.columns and new_r not in df.columns:
            df[new_r] = df[src_r]
            added.append(new_r)
    if added:
        if not meta_line.endswith(("\n", "\r")):
            meta_line += newline
        with open(ts_csv_path, "w", newline="", encoding="utf-8") as f:
            f.write(meta_line)
            _to_csv(df, f, newline)
    return added


# ----------------------------------------------------------------------
# Set_filter_file.xlsx selection sheets
# ----------------------------------------------------------------------
def append_selection_rows(xlsx_path, sheet_name, members, selected=0):
    """
    Append members to a selection sheet *without altering existing rows*.

    Used to keep the universal Set_filter_file.xlsx complete (every choice
    present) while leaving its 0/1 selections untouched. Returns added members.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]
    present = set()
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is not None:
            present.add(str(name))
    next_row = ws.max_row + 1
    added = []
    for m in members:
        if str(m) in present:
            continue
        ws.cell(row=next_row, column=1).value = m
        ws.cell(row=next_row, column=2).value = selected
        next_row += 1
        added.append(m)
    wb.save(xlsx_path)
    return added


def update_selection_sheet(xlsx_path, sheet_name, enabled, add_rows=()):
    """
    Edit a selection sheet of Set_filter_file.xlsx.

    Column A holds set members, column B the 0/1 enable flag.

      enabled  : iterable of member names to set to 1; every other existing
                 member is set to 0.
      add_rows : member names to append if not already present (appended
                 enabled iff also in ``enabled``).
    """
    enabled = set(str(e) for e in enabled)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]

    present = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is None:
            continue
        present[str(name)] = r
        ws.cell(row=r, column=2).value = 1 if str(name) in enabled else 0

    next_row = ws.max_row + 1
    for member in add_rows:
        if str(member) in present:
            continue
        ws.cell(row=next_row, column=1).value = member
        ws.cell(row=next_row, column=2).value = 1 if str(member) in enabled else 0
        next_row += 1

    wb.save(xlsx_path)


# ----------------------------------------------------------------------
# High-level driver helpers
# ----------------------------------------------------------------------
def expand_all_parameters(parameters_dir, region_map=None, target_years=None,
                          verbose=True):
    """
    Apply region duplication and/or year interpolation to every master
    parameter CSV. Either step is skipped when its argument is None.
    """
    for csv in param_csv_paths(parameters_dir):
        df = read_param(csv)
        before = len(df)
        if region_map:
            df = duplicate_regions_in_df(df, region_map)
        if target_years:
            df = interpolate_years_in_df(df, target_years)
        if len(df) != before:
            write_param(df, csv)
            if verbose:
                print(f"  {os.path.basename(csv):42s} {before:7d} -> {len(df):7d} rows")
        elif verbose:
            print(f"  {os.path.basename(csv):42s} unchanged")
