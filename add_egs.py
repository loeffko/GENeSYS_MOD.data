# -*- coding: utf-8 -*-
"""Add four EGS (Enhanced Geothermal Systems) technologies P_EGS_R1..R4 to the
GENeSYS-MOD data, sourced from `NA_inputs/EGS_LCOE_model.xlsx` and
`NA_inputs/geothermal_allocation_model.xlsx`.

Inserted CSV rows:
  - Sets_Technology: add P_EGS_R1..R4
  - Par_CapitalCost: Central-scenario CapEx per region (R1..R4), 2025-2040
  - Par_FixedCost: 130 MEUR/GW-yr, World, All years
  - Par_OperationalLife: 30 yr
  - Par_AvailabilityFactor: 0.9 (90% CF, World, All)
  - Par_CapacityToActivityUnit: 31.536 (PJ/GW-yr)
  - Par_OutputActivityRatio: Power, mode 1, All, 1
  - Par_TagTechnologyToSubsets: subset "EGS" + standard subsets
  - Par_TagTechnologyToSector: Power
  - Par_TotalAnnualMaxCapacity: per ISO × P_EGS_R*, from Region x ISO crosswalk
  - Par_GroupTotalAnnualMinCapacity: USA × EGS × year (Low scenario)
  - Par_GroupTotalAnnualMaxCapacity: USA × EGS × year (Central scenario)
  - Par_ResidualCapacity: 2025 baseline, linear decline to 0 by 2040
                          (P_EGS_R1 only — existing fleet is conventional)
"""
import csv
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
NA_DIR = os.path.join(HERE, "NA_inputs")
PAR_DIR = os.path.join(HERE, "Data", "Parameters")
SETS_DIR = os.path.join(PAR_DIR, "00_Sets&Tags")

TECHS = ["P_EGS_R1", "P_EGS_R2", "P_EGS_R3", "P_EGS_R4"]

# ISO from EGS sheet -> model region name
ISO_MAP = {
    "CAISO": "California",  "WECC": "WECC",  "ERCOT": "ERCOT",
    "SPP": "SPP",           "MISO": "MISO",  "SERC": "SERC",
    "PJM": "PJM",           "ISO-NE": "NewEngland", "NYISO": "NewYork",
}

STAMP = "2026-06-04"
AUTHOR = "Konstantin Loffler <kl@wip.tu-berlin.de>"
SRC_LCOE = "EGS_LCOE_model.xlsx (Central scenario, ATB 2024) [USD->EUR 0.92]"
SRC_ALLOC = "geothermal_allocation_model.xlsx (Sebastian, April 2026)"

# Source data in 2022$. ECB 2022 average EUR/USD ~0.95; using 0.92 as a
# round-ish forward-looking rate consistent with GENeSYS-MOD conventions.
USD_TO_EUR = 0.92


# ---------- read source data ----------
def read_lcoe():
    wb = openpyxl.load_workbook(os.path.join(NA_DIR, "EGS_LCOE_model.xlsx"),
                                read_only=True, data_only=True)
    # Region x ISO crosswalk (rows 12-16, cols A..J)
    ws = wb["Region_Definitions"]
    rows = list(ws.iter_rows(values_only=True))
    iso_hdr = rows[12]   # ['Region','WECC','ERCOT','SPP','MISO','CAISO','SERC','PJM','ISO-NE','NYISO']
    max_cap = {}  # (tech, iso_model_region) -> GW
    for r in rows[13:17]:
        region_code = r[0]   # R1..R4
        tech = f"P_EGS_{region_code}"
        for j, iso in enumerate(iso_hdr[1:], start=1):
            val = r[j]
            model_reg = ISO_MAP.get(iso)
            if model_reg and val is not None and val > 0:
                max_cap[(tech, model_reg)] = float(val)

    # Central CapEx 2025-2040 per region (rows 18-22)
    ws = wb["LCOE_Components"]
    rows = list(ws.iter_rows(values_only=True))
    central_yrs = rows[17][1:17]  # 16 years
    capex = {}  # (tech, year) -> $/kW
    for off, code in enumerate(["R1", "R2", "R3", "R4"]):
        tech = f"P_EGS_{code}"
        row_vals = rows[19 + off][1:17]
        for y, v in zip(central_yrs, row_vals):
            capex[(tech, int(y))] = float(v)
    return max_cap, capex


def read_allocation():
    """Everything comes from the Output sheet:

    - baseline: per-ISO Low-scenario 2025 values (rows 5-13) — used as the
      model's 2025 residual fleet. Sums to the CONUS total 2025 (~4.07 GW),
      so the base-year residual sits exactly on the group-min trajectory.
    - low / central: the "CONUS total" rows (NOT the headline rows — those
      carry a +0.07 GW non-CONUS offset that no ISO region can supply, which
      made GroupMin 2025 (4.14) exceed the residual fleet => infeasible).
    """
    wb = openpyxl.load_workbook(os.path.join(NA_DIR, "geothermal_allocation_model.xlsx"),
                                read_only=True, data_only=True)
    ws = wb["Output"]
    rows = list(ws.iter_rows(values_only=True))

    def conus_series(header_idx, conus_idx):
        years, vals = rows[header_idx], rows[conus_idx]
        assert str(vals[0]).startswith("CONUS"), f"expected CONUS row at {conus_idx}, got {vals[0]!r}"
        return {int(y): float(v) for y, v in zip(years[1:], vals[1:])
                if y is not None and int(y) >= 2025}

    low     = conus_series(4, 14)    # Low block
    central = conus_series(19, 29)   # Central block

    # Per-ISO Low 2025 (Low block rows 5-13; column 2 = year 2025)
    baseline = {}
    for r in rows[5:14]:
        iso, gw = r[0], r[2]
        mr = ISO_MAP.get(iso)
        if mr and gw is not None and float(gw) > 0:
            baseline[mr] = float(gw)
    return baseline, low, central


# ---------- CSV helpers ----------
def append_rows(path, rows):
    with open(path, "a", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        for r in rows:
            w.writerow(r)


def existing_techs_in(path, col=0):
    with open(path, "r", encoding="utf-8") as f:
        next(f, None)
        return {line.split(",")[col].strip() for line in f if line.strip()}


# ---------- per-target updaters ----------
def update_sets_technology():
    path = os.path.join(SETS_DIR, "Sets_Technology.csv")
    have = existing_techs_in(path, col=0)
    new = [t for t in TECHS if t not in have]
    if new:
        with open(path, "a", encoding="utf-8", newline="") as f:
            for t in new:
                f.write(t + "\n")
        print(f"  Sets_Technology: +{len(new)}")
    else:
        print("  Sets_Technology: already present")


def update_capital_cost(capex):
    path = os.path.join(PAR_DIR, "Par_CapitalCost", "Par_CapitalCost.csv")
    rows = [["World", t, str(y), str(round(v * USD_TO_EUR, 2)), "", "MEUR/GW",
             SRC_LCOE, STAMP, AUTHOR]
            for (t, y), v in sorted(capex.items())]
    _drop_tech_rows(path, TECHS, tech_col=1)
    append_rows(path, rows)
    print(f"  Par_CapitalCost: +{len(rows)} (USD->EUR x{USD_TO_EUR})")


def update_fixed_cost():
    path = os.path.join(PAR_DIR, "Par_FixedCost", "Par_FixedCost.csv")
    _drop_tech_rows(path, TECHS, tech_col=1)
    fom = round(130 * USD_TO_EUR, 2)  # 130 $/kW-yr -> EUR-equivalent
    rows = [["World", t, "All", str(fom), "", "MEUR/GW", SRC_LCOE, STAMP, AUTHOR, ""]
            for t in TECHS]
    append_rows(path, rows)
    print(f"  Par_FixedCost: +{len(rows)} (FOM = {fom} MEUR/GW-yr)")


def update_operational_life():
    path = os.path.join(PAR_DIR, "Par_OperationalLife", "Par_OperationalLife.csv")
    _drop_tech_rows(path, TECHS, tech_col=0)
    rows = [[t, "30", "", "Years", SRC_LCOE, STAMP, AUTHOR] for t in TECHS]
    append_rows(path, rows)
    print(f"  Par_OperationalLife: +{len(rows)}")


def update_availability_factor():
    path = os.path.join(PAR_DIR, "Par_AvailabilityFactor", "Par_AvailabilityFactor.csv")
    _drop_tech_rows(path, TECHS, tech_col=1)
    rows = [["World", t, "All", "0.9", "", "Fraction", SRC_LCOE, STAMP, AUTHOR] for t in TECHS]
    append_rows(path, rows)
    print(f"  Par_AvailabilityFactor: +{len(rows)}")


def update_capacity_to_activity():
    path = os.path.join(SETS_DIR, "Par_CapacityToActivityUnit.csv")
    _drop_tech_rows(path, TECHS, tech_col=0)
    # Header is exactly 2 cols: Technology,Value
    rows = [[t, "31.536"] for t in TECHS]
    append_rows(path, rows)
    print(f"  Par_CapacityToActivityUnit: +{len(rows)}")


def update_oar():
    path = os.path.join(PAR_DIR, "Par_OutputActivityRatio", "Par_OutputActivityRatio.csv")
    _drop_tech_rows(path, TECHS, tech_col=1)
    rows = [["World", t, "Power", "1", "All", "1", "", "PJ",
             "EGS dispatchable geothermal", STAMP, AUTHOR] for t in TECHS]
    append_rows(path, rows)
    print(f"  Par_OutputActivityRatio: +{len(rows)}")


def update_tag_subsets():
    path = os.path.join(SETS_DIR, "Par_TagTechnologyToSubsets.csv")
    _drop_tech_rows(path, TECHS, tech_col=0)
    rows = []
    for t in TECHS:
        rows.append([t, "EGS", "1", "", "Binary", "not relevant", STAMP, AUTHOR])
        rows.append([t, "Renewables", "1", "", "Binary", "not relevant", STAMP, AUTHOR])
        rows.append([t, "PowerSupply", "1", "", "Binary", "not relevant", STAMP, AUTHOR])
    append_rows(path, rows)
    print(f"  Par_TagTechnologyToSubsets: +{len(rows)}")


def update_tag_sector():
    path = os.path.join(SETS_DIR, "Par_TagTechnologyToSector.csv")
    _drop_tech_rows(path, TECHS, tech_col=0)
    # Header is exactly 3 cols: Technology,Sector,Value
    rows = [[t, "Power", "1"] for t in TECHS]
    append_rows(path, rows)
    print(f"  Par_TagTechnologyToSector: +{len(rows)}")


def update_max_capacity(max_cap):
    path = os.path.join(PAR_DIR, "Par_TotalAnnualMaxCapacity", "Par_TotalAnnualMaxCapacity.csv")
    _drop_tech_rows(path, TECHS, tech_col=1)
    rows = [[r, t, "All", str(round(gw, 3)), "", "GW", SRC_LCOE, STAMP, AUTHOR]
            for (t, r), gw in sorted(max_cap.items())]
    append_rows(path, rows)
    print(f"  Par_TotalAnnualMaxCapacity: +{len(rows)}")


def update_group_caps(low, central):
    # Min = Low scenario; Max = Central. Year subset = 2025-2040 model years.
    # Min is FLOORED and Max CEILED at 3 decimals so per-row rounding of the
    # residual capacities can never push the base-year fleet outside the cone.
    import math
    yrs = sorted({y for y in low.keys() if 2025 <= y <= 2050})
    rows_min = [["EGS", "USA", str(y), str(math.floor(low[y] * 1000) / 1000), "", "GW", SRC_ALLOC, STAMP, AUTHOR]
                for y in yrs]
    rows_max = [["EGS", "USA", str(y), str(math.ceil(central[y] * 1000) / 1000), "", "GW", SRC_ALLOC, STAMP, AUTHOR]
                for y in yrs]
    p_min = os.path.join(PAR_DIR, "Par_GroupTotalAnnualMinCapacity", "Par_GroupTotalAnnualMinCapacity.csv")
    p_max = os.path.join(PAR_DIR, "Par_GroupTotalAnnualMaxCapacity", "Par_GroupTotalAnnualMaxCapacity.csv")
    _drop_subset_rows(p_min, "EGS", "USA")
    _drop_subset_rows(p_max, "EGS", "USA")
    append_rows(p_min, rows_min)
    append_rows(p_max, rows_max)
    print(f"  GroupTotalAnnualMinCapacity (Low): +{len(rows_min)}")
    print(f"  GroupTotalAnnualMaxCapacity (Central): +{len(rows_max)}")


# EGS build-rate smoothing: SC2's %-of-max cap is toothless for EGS (its max is the
# full resource potential ~7000 GW), so the build path is smoothed with the new
# GroupTotalAnnualMaxNewCapacity parameter (caps summed annual NewCapacity over a
# tech subset × region subset; TCC5 in genesysmod_equ.jl). Region-specific via a
# single-region subset, so no per-region hardcode lives in the model code.
EGS_NEWCAP_LIMIT = {"ERCOT": 2.4}   # GW/yr of new EGS capacity, per region


def update_group_new_caps():
    yrs = list(range(2025, 2041))
    # 1) ensure a single-region region-subset exists for each capped region
    tag_path = os.path.join(SETS_DIR, "Par_TagRegionToSubsets.csv")
    have = set()
    with open(tag_path, "r", encoding="utf-8") as f:
        next(f, None)
        for line in f:
            p = line.split(",")
            if len(p) >= 2:
                have.add((p[0].strip(), p[1].strip()))
    sub_rows = [[reg, reg, "1", "", "Binary", "not relevant", STAMP, AUTHOR]
                for reg in EGS_NEWCAP_LIMIT if (reg, reg) not in have]
    if sub_rows:
        append_rows(tag_path, sub_rows)
        print(f"  Par_TagRegionToSubsets (single-region subsets): +{len(sub_rows)}")
    # 2) create the new param dir+file (header) if missing, then write EGS rows
    d = os.path.join(PAR_DIR, "Par_GroupTotalAnnualMaxNewCapacity")
    path = os.path.join(d, "Par_GroupTotalAnnualMaxNewCapacity.csv")
    if not os.path.exists(path):
        os.makedirs(d, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="") as f:
            f.write("TechnologySubset,RegionSubset,Year,Value,,Unit,Source,Updated at,Updated by\n")
    rows = []
    for reg, lim in EGS_NEWCAP_LIMIT.items():
        _drop_subset_rows(path, "EGS", reg)
        rows += [["EGS", reg, str(y), str(lim), "", "GW",
                  "EGS build-rate smoothing (TU Berlin assumption)", STAMP, AUTHOR] for y in yrs]
    append_rows(path, rows)
    print(f"  Par_GroupTotalAnnualMaxNewCapacity (EGS x {list(EGS_NEWCAP_LIMIT)}): +{len(rows)}")


def update_residual_capacity(baseline):
    # 2025 = baseline (P_EGS_R1 only — existing fleet best represented as Prime).
    # Linear decline to 0 by 2040 (avg fleet age ~30yr in 2025; 30yr lifetime).
    # Emit every year 2025-2040 explicitly so the model does not interpret a
    # missing year as zero residual capacity.
    path = os.path.join(PAR_DIR, "Par_ResidualCapacity", "Par_ResidualCapacity.csv")
    _drop_tech_rows(path, TECHS, tech_col=1)
    rows = []
    for reg, gw in sorted(baseline.items()):
        for yr in range(2025, 2041):
            frac = max(0.0, 1.0 - (yr - 2025) / 15.0)   # 1 at 2025, 0 at 2040
            rows.append([reg, "P_EGS_R1", str(yr), str(round(gw * frac, 4)), "", "GW",
                         "EIA 2025 Geothermal Market Report; linear decline 2025-2040 (avg fleet age 30yr, 30yr life)",
                         STAMP, AUTHOR])
    append_rows(path, rows)
    print(f"  Par_ResidualCapacity (P_EGS_R1, every year 2025-2040): +{len(rows)}")


# ---------- low-level CSV row drop ----------
def _drop_tech_rows(path, techs, tech_col):
    with open(path, "r", encoding="utf-8") as f:
        lines = f.readlines()
    kept = []
    for i, line in enumerate(lines):
        if i == 0:
            kept.append(line); continue
        parts = line.split(",")
        if len(parts) > tech_col and parts[tech_col].strip() in techs:
            continue
        kept.append(line)
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.writelines(kept)


def _drop_subset_rows(path, tech_subset, region_subset):
    """Drop rows where (col0, col1) == (tech_subset, region_subset)."""
    with open(path, "r", encoding="utf-8") as f:
        lines = f.readlines()
    kept = []
    for i, line in enumerate(lines):
        if i == 0:
            kept.append(line); continue
        parts = line.split(",")
        if len(parts) >= 2 and parts[0].strip() == tech_subset and parts[1].strip() == region_subset:
            continue
        kept.append(line)
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.writelines(kept)


def main():
    print("Loading source data...")
    max_cap, capex = read_lcoe()
    baseline, low, central = read_allocation()
    print(f"  max_cap entries: {len(max_cap)}")
    print(f"  capex entries:   {len(capex)}")
    print(f"  baseline ISOs:   {sorted(baseline.keys())}")
    print(f"  trajectory years (low/central): {sorted(low.keys())}")

    print("\nUpdating CSVs...")
    update_sets_technology()
    update_capital_cost(capex)
    update_fixed_cost()
    update_operational_life()
    update_availability_factor()
    update_capacity_to_activity()
    update_oar()
    update_tag_subsets()
    update_tag_sector()
    update_max_capacity(max_cap)
    update_group_caps(low, central)
    update_group_new_caps()
    update_residual_capacity(baseline)
    print("\nDone.")


if __name__ == "__main__":
    main()
