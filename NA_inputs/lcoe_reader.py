# -*- coding: utf-8 -*-
"""Reader for the FEL2026 LCOE cost model workbook (v1.4, dated 2026-06-18).

Single source of truth for transferring LCOE-model cost data into the
parameter CSVs (NA regional rows via NA_inputs/add_lcoe_costs.py, EU World
rows via apply_lcoe_eu_costs.py). Reads planned CAPEX, Fixed OPEX and WACC
from Source_Input_Wide and the real/planned multiplier from CAPEX_Config.

Known LCOE-file issues handled here (documented in
LCOE_Model_v1.4_issues_2026-07-23.md):
  - stray 0.5 CAPEX multipliers in 2025/26 (wind/SOFC/SMR/Solar-PV rows in
    CAPEX_Config): IGNORED — only the PV-utility learning-curve multiplier
    (0.98 -> 0.724, monotone) is applied; all other multipliers are treated
    as 1 (planned CAPEX).
  - nuclear 'real/planned' multiplier (x3.0 US / x2.75 EU) double-counts
    market escalation: nuclear/SMR are NOT imported at all.
  - geothermal rows describe a different asset class (13-19 kEUR/kW) than the
    model's conventional/EGS technologies: not imported.
  - Recip-engine data only exists on the 'Customized_LCOE SOFC US' sheet
    (CAPEX 3125 EUR/kW, FOM 29 EUR/kW/yr); efficiency there (0.64) is a CCGT
    copy-bug and is not imported.
"""
import openpyxl

SOURCE_LABEL = "FEL2026 LCOE Model file V1.4, dated 2026-06-18"

def load(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"series": {}, "mult": {}}
    ws = wb["Source_Input_Wide"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    ycols = {i: int(v) for i, v in enumerate(rows[0]) if str(v).isdigit()}
    for r in rows[1:]:
        if not r or r[1] not in ("USA", "All regions"):
            continue
        if r[3] not in ("No CCS", "Not applicable"):
            continue
        key = (r[1], r[2], r[4])          # (country, tech, kpi)
        vals = {y: float(r[i]) for i, y in ycols.items()
                if isinstance(r[i], (int, float))}
        if vals:
            out["series"][key] = vals
    ws2 = wb["CAPEX_Config"]
    rows2 = list(ws2.iter_rows(min_row=5, values_only=True))
    yc2 = {i: int(v) for i, v in enumerate(rows2[0]) if str(v).isdigit()}
    for r in rows2[1:]:
        if not r or r[2] not in ("USA", "All regions"):
            continue
        key = (r[2], r[3])
        vals = {y: float(r[i]) for i, y in yc2.items()
                if isinstance(r[i], (int, float))}
        if vals and key not in out["mult"]:
            out["mult"][key] = vals
    wb.close()
    return out

def series(data, country, tech, kpi, years):
    """Yearly values for `years`; nearest-known-year fallback at the edges."""
    s = data["series"].get((country, tech, kpi))
    if not s:
        return None
    known = sorted(s)
    def at(y):
        if y in s:
            return s[y]
        lo = [k for k in known if k < y]
        hi = [k for k in known if k > y]
        if lo and hi:                      # linear interpolation
            a, b = lo[-1], hi[0]
            return s[a] + (s[b] - s[a]) * (y - a) / (b - a)
        return s[lo[-1]] if lo else s[hi[0]]
    return {y: at(y) for y in years}

def pv_learning(data, country, years):
    """PV-utility learning multiplier (the only CAPEX_Config row applied)."""
    m = data["mult"].get((country, "PV: Utility scale"), {})
    return {y: m.get(y, 1.0) for y in years}
