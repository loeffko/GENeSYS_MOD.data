# -*- coding: utf-8 -*-
"""Capacity bounds + RES potentials (merged).

Writes the following NA-side parameters in one pass so guardrail vs. potential
do not overlap:

  Par_ResidualCapacity         : 2025 base * (1 - 0.05*(y-2025)), 2025-2040 —
                                 LINEAR retirement of 5% of the 2025 fleet per
                                 year (constant absolute). P_Nuclear is held
                                 flat (rate 0) — fleet assumed to stay on.
                                 EXCEPTIONS: CCGT+OCGT follow the national
                                 announced-retirement schedule GAS_RETIREMENT_GW
                                 (~26 GW to 2035, ramping to 13 GW/yr by 2040,
                                 allocated by 2025 fleet share); P_Gas_Steam
                                 follows the capacity file's own per-region ST
                                 decline (trend-extended past 2035). Gas techs
                                 additionally get a monotonic (never-decreasing)
                                 TotalAnnualMaxCapacity.
                                 Hydro is the exception: split reservoir/RoR,
                                 residual held FLAT (no retirement) with growth
                                 forced via a strict TotalAnnualMinCapacity (see
                                 the hydro block + Par_ResidualCapacity/
                                 Assumptions.txt). Storage is split into PHS
                                 (fixed existing fleet) + Li-Ion BESS (residual +
                                 min forced to the US Pools storage trajectory,
                                 max open) — block 1d.

  Par_TotalAnnualMinCapacity   : guardrail funnel 2026-2040
      2026-2028:  min = val*0.98 (±2%)
      2029-2035:  linearly widen to min = val*0.90 at 2035
      2036-2040:  hold the 2035 funnel min (no contraction post-2035)
      Nuclear is pinned to 1.0 (no downward widening).

  Par_TotalAnnualMaxCapacity   : guardrail funnel 2026-2035 widening up to 1.30
      For PV_Utility_Opt / Wind_Onshore_Opt (guardrail reps that have a
      restool potential): 2036-2040 linearly interp from the 2035 funnel
      max -> the per-region restool potential at 2040, every year filled.
      For Nuclear/Gas/Hydro (no restool potential): hold 2035 funnel max
      through 2040.
      Overflow variants (PV_Utility_Avg/Inf, Wind_Onshore_Avg/Inf,
      A_Rooftop_*) and Canada: every year 2025-2040 = restool potential
      (flat, no growth, since the value is the physical ceiling). The rep
      (_Opt) is filled first; capacity beyond its share-of-potential cascades
      into _Avg then _Inf, so the observed fleet sits on the high-CF Opt class.

  Par_TagTechnologyToSubsets   : P_Nuclear -> Nuclear subset (idempotent)
  Par_GroupTotalAnnualMinCapacity : Nuclear x USA target trajectory 2035-2040

Sources:
  - NA_inputs/US Pools - Generation and Capacity_anonymized.xlsx (Pool-Region x
    Fuel Group x year 2025-2035, Capacity MW). Anonymized; values clamped >=0.
  - NA_restool/northamerica_potentials_combined.csv (PV/Wind/Rooftop GW)
  - NA_restool/canada_potentials_combined.csv

Idempotent: existing NA rows for the managed techs are removed from each CSV
before append. Canada bounds come only from the restool (Canada is not in the
US Pools file).

Run:  python NA_inputs/add_capacity_bounds.py            # dry-run (sample print)
      python NA_inputs/add_capacity_bounds.py --apply
"""
import os, sys
import pandas as pd
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_REPO = os.path.normpath(os.path.join(HERE, ".."))
SRC = os.path.join(HERE, "US Pools - Generation and Capacity_v2.xlsx")
# US nuclear forecast (per-unit; Summary "Total LSR") + US coal trajectory model
NUCLEAR_SRC = os.path.join(HERE, "PMK_Nuclear Operating and Forecast Through 2040_4.15.2026.xlsx")
COAL_SRC = os.path.join(HERE, "Coal_Trajectory_Model.xlsx")
POT_NA = os.path.join(DATA_REPO, "NA_restool", "northamerica_potentials_combined.csv")
POT_CA = os.path.join(DATA_REPO, "NA_restool", "canada_potentials_combined.csv")
PARAM = lambda n: os.path.join(DATA_REPO, "Data", "Parameters", n, n + ".csv")
apply = "--apply" in sys.argv
# Sensitivity options:
#   --fel <file.xlsx>   demand workbook for the funnel demand/generation ratio
#                       (default base FEL; must match convert_fel_to_demand --fel)
#   --max-upscale       let a demand/generation ratio > 1 RAISE the funnel max
#                       (dc_high sensitivity); the min stays capped at 1.0
#   --funnel economic   wider funnel from ~2030 (min x0.75, max x1.5 by 2035):
#                       the model decides more, data steers less
#   --funnel grid       slow-start accelerating widening 2030->2040 (quadratic;
#                       min x0.70 / max x1.50 at 2040) - grid_high
MAX_UPSCALE = "--max-upscale" in sys.argv
# Demand upscaling hits RES ceilings at only HALF strength (dc_high family):
# the boom buys firm capacity headroom fully, RES potential-derived ceilings
# rise by half the demand ratio's excess (guardrail reps; the overflow classes
# scale from the rep automatically).
RES_HALF_UPSCALE = {"P_PV_Utility_Opt", "P_Wind_Onshore_Opt"}
#   --scenario-subdir <name>  write outputs into the conversion's scenario
#                             subfolders (Par_X/<name>/Par_X.csv, row-upsert over
#                             the base at conversion time) instead of mutating
#                             the base CSVs - the sensitivity workflow.
SCENARIO_SUBDIR = sys.argv[sys.argv.index("--scenario-subdir") + 1] if "--scenario-subdir" in sys.argv else None
#   --gas-min-floor <x>  override the 0.93 gas blend floor (recession: 0 = the
#                        demand ratio cuts the gas floor freely post-2030)
if "--gas-min-floor" in sys.argv:
    GAS_MIN_BLEND_FLOOR_OVERRIDE = float(sys.argv[sys.argv.index("--gas-min-floor") + 1])
else:
    GAS_MIN_BLEND_FLOOR_OVERRIDE = None
FUNNEL_STYLE = sys.argv[sys.argv.index("--funnel") + 1] if "--funnel" in sys.argv else "base"
#   --bess-min-relax     (bess_pessimistic) relax the Li-Ion BESS min post-2030:
#                        the forced US-Pools trajectory blends linearly down to
#                        x0.75 of itself at 2040, so the market may UNDERDELIVER
#                        the storage outlook (the near-term pipeline stays firm).
#   --bess-min-relax <f2040>  (bess_pessimistic) blend the forced Li-Ion min
#                        linearly from x1.0 at 2030 to x<f2040> at 2040
#                        (0.53 matches the agreed ~147 GW low build-out)
BESS_MIN_RELAX_2040 = (float(sys.argv[sys.argv.index("--bess-min-relax") + 1])
                       if "--bess-min-relax" in sys.argv else None)
#   --bess-pin           with --bess-min-relax: the Li-Ion MAX follows the
#                        relaxed min (hard build-out pin, no economic upside)
BESS_PIN = "--bess-pin" in sys.argv
def bess_min_relax_f(y):
    if BESS_MIN_RELAX_2040 is None:
        return 1.0
    return 1.0 + (BESS_MIN_RELAX_2040 - 1.0) * max(0.0, min(1.0, (y - 2030) / 10.0))
#   --gas-min-relax <f2040>  (bess_cost_low_8h) SLIGHT gas-min opening: the gas
#                        floor blends linearly from x1.0 at 2030 to x<f2040> at
#                        2040 (e.g. 0.9 - much gentler than the grid funnel's
#                        0.70 or the economic funnel's 0.75), so extreme cheap
#                        storage may substitute a little firm gas capacity.
GAS_MIN_RELAX_2040 = (float(sys.argv[sys.argv.index("--gas-min-relax") + 1])
                      if "--gas-min-relax" in sys.argv else None)
def gas_min_relax_f(y):
    if GAS_MIN_RELAX_2040 is None:
        return 1.0
    return 1.0 + (GAS_MIN_RELAX_2040 - 1.0) * max(0.0, min(1.0, (y - 2030) / 10.0))
#   --max-boost <x>      scale the funnel MAX a further x for the main expansion
#                        techs (PV, onshore wind, Li-Ion BESS) on top of any
#                        demand upscaling (dc_high_limitless)
#   --max-boost-regions <a,b>  restrict the boost to these regions (default all;
#                        dc_high_limitless boosts ERCOT only - elsewhere the
#                        demand shift re-routes gas builds instead)
MAX_BOOST = float(sys.argv[sys.argv.index("--max-boost") + 1]) if "--max-boost" in sys.argv else None
MAX_BOOST_PREFIXES = ("P_PV_", "P_Wind_Onshore", "D_Battery_Li-Ion")
MAX_BOOST_REGIONS = (set(sys.argv[sys.argv.index("--max-boost-regions") + 1].split(","))
                     if "--max-boost-regions" in sys.argv else None)
ECON_MIN_EXTRA_2035, ECON_MAX_EXTRA_2035 = 0.75, 1.50
#   --funnel grid    (grid_high) widening starts 2030 SLOWLY and accelerates
#                    towards 2040 (quadratic ramp; keeps widening post-2035):
#                    min x0.925 / max x1.125 at 2035 -> min x0.70 / max x1.50
#                    at 2040. More grid should displace gas with remote RES;
#                    the base funnel would pin the mix and hide that response.
GRID_MIN_EXTRA_2040, GRID_MAX_EXTRA_2040 = 0.70, 1.50
def econ_min_f(y):
    if FUNNEL_STYLE == "economic":
        return 1.0 + (ECON_MIN_EXTRA_2035 - 1.0) * max(0.0, min(1.0, (y - 2030) / 5.0))
    if FUNNEL_STYLE == "grid":
        return 1.0 + (GRID_MIN_EXTRA_2040 - 1.0) * max(0.0, min(1.0, (y - 2030) / 10.0)) ** 2
    return 1.0
def econ_max_f(y):
    if FUNNEL_STYLE == "economic":
        return 1.0 + (ECON_MAX_EXTRA_2035 - 1.0) * max(0.0, min(1.0, (y - 2030) / 5.0))
    if FUNNEL_STYLE == "grid":
        return 1.0 + (GRID_MAX_EXTRA_2040 - 1.0) * max(0.0, min(1.0, (y - 2030) / 10.0)) ** 2
    return 1.0
# post-2035 continuation of the style ramp: the year loop derives 2036-2040
# values from the 2035 anchors (which already carry the 2035 multiplier), so a
# ramp that keeps moving after 2035 (grid) re-scales them by the RATIO to its
# 2035 level. base/economic ramps are flat post-2035 -> ratio 1 (no change).
def post35_min_f(y):
    return econ_min_f(y) / econ_min_f(2035)
def post35_max_f(y):
    return econ_max_f(y) / econ_max_f(2035)

# Guardrail category -> representative model tech (US Pools dataset,
# "Fuel + Technology" split since the 2026-06 file update).
# "Coal" is handled separately (lignite/hardcoal split below).
# "Storage" is split into PHS + Li-Ion BESS in a dedicated block (1d).
TECH = {"Natural Gas - CCCT":  "P_Gas_CCGT",
        "Natural Gas - SCCT":  "P_Gas_OCGT",
        "Natural Gas - ST":    "P_Gas_Steam",
        "Natural Gas - Other": "P_Gas_Engines",
        "Solar": "P_PV_Utility_Opt",
        "Wind": "P_Wind_Onshore_Opt",
        "Nuclear": "P_Nuclear"}
MODEL_TECHS_GUARDRAIL = set(TECH.values())

# Hydro is split into reservoir + run-of-river and handled in a dedicated block
# (NOT the generic decaying-residual loop): the existing fleet does not retire,
# so ResidualCapacity is FLAT at the 2025 value, and growth to the planned
# trajectory is forced through TotalAnnualMinCapacity (strict to the US Pools
# capacity sheet, with the 2030-2035 trend extrapolated to 2040). Per-region
# run-of-river share — methodology + sources in
# Data/Parameters/Par_ResidualCapacity/Assumptions.txt.
HYDRO_CATEGORY = "Hydro"
HYDRO_TECHS = {"P_Hydro_Reservoir", "P_Hydro_RoR"}
HYDRO_ROR_SHARE = {"NewYork": 0.80, "WECC": 0.60, "MISO": 0.60, "SPP": 0.60,
                   "NewEngland": 0.50, "California": 0.35, "SERC": 0.20,
                   "ERCOT": 0.15, "PJM": 0.10, "Canada": 0.30}
HYDRO_ROR_DEFAULT = 0.40
HYDRO_MAX_HEADROOM = 1.10   # TotalAnnualMaxCapacity = strict sheet min * 1.10

# Storage split: the US Pools "Storage" category lumps pumped hydro + batteries.
# It is split into existing pumped hydro (D_PHS) and battery storage, which is
# mapped to Li-Ion (essentially all deployed US grid batteries today are
# Li-Ion). The 2025 PHS fleet per region (GW, EIA-860 / known plants — see
# Assumptions.txt); the remainder of US Pools "Storage" (and all of its growth)
# is Li-Ion BESS. PHS existing fleet persists (residual = min = 2025 value);
# its max is not a hard pin but a growth-rate-capped ceiling — the real-world
# limit on new US PHS is the buildout rate (long lead times, siting), not a
# hard cap. PHS_MAX_GROWTH lets the fleet grow at most that fraction/yr
# (compound) from 2025; regions with no 2025 PHS stay ~0 (no greenfield).
STORAGE_CATEGORY = "Storage"
STORAGE_TECHS = {"D_PHS", "D_Battery_Li-Ion", "D_CAES"}
STORAGE_UNCAPPED = 999999   # Li-Ion BESS max left open above the planned floor
# Redox flow batteries (D_Battery_Redox): emerging tech, no existing fleet. Forbid new
# build before REDOX_START_YEAR, then cap annual additions to REDOX_ANNUAL_CAP GW/region
# via a ramping TotalAnnualMaxCapacity — otherwise the optimiser builds tens of GW in the
# first year (e.g. >50 GW ERCOT 2026).
REDOX_TECHS = {"D_Battery_Redox"}
REDOX_START_YEAR = 2030
REDOX_ANNUAL_CAP = 2.0      # GW/yr per region, allowed only from REDOX_START_YEAR
PHS_MAX_GROWTH = 0.015      # D_PHS max: <=1.5%/yr compound on the 2025 fleet
                            # (~22.8 -> ~28.5 GW NA by 2040, i.e. +~6 GW new,
                            # matching the realistic US pipeline: Goldendale,
                            # Gordon Butte, Eagle Mountain, Swan Lake, ...)
# CAES (D_CAES): salt-cavern / A-CAES buildout is deployment-limited, not
# geology-limited (CAES has deployed ~0.4 GW worldwide in 45 yr). Cap the
# optimiser at a per-region 2040 ceiling (salt-weighted: Gulf Coast/Permian ->
# ERCOT/SERC/SPP, Michigan/Williston -> MISO, Appalachian -> PJM/NewYork, A-CAES
# -> WECC/California/Canada; NewEngland ~0, no geology). Optimistic ~12 GW NA
# total, ramped linearly from 0 in 2025. Without it the LP over-builds CAES as a
# generic cheap long-duration store.
CAES_MAX_GW_2040 = {"ERCOT": 3.0, "SERC": 2.0, "SPP": 1.5, "MISO": 1.5,
                    "PJM": 1.0, "WECC": 1.0, "California": 0.5, "NewYork": 0.5,
                    "Canada": 1.0, "NewEngland": 0.0}
# A TotalAnnualMaxCapacity of exactly 0 is read as "unset" and converted to
# 999999 (= no limit) by genesysmod_bounds.jl for storage/transformation techs.
# To actually FORBID / cap at ~0 (e.g. PHS in ERCOT/SPP, CAES in NewEngland, the
# 2025 CAES ramp start) we must write a tiny nonzero ceiling instead of 0.
FORBID_EPS = 0.001  # GW (~1 MW): effectively zero but survives the 0->999999 rule
PHS_GW_2025 = {"California": 4.0, "SERC": 6.5, "PJM": 5.0, "MISO": 2.6,
               "NewEngland": 1.8, "WECC": 1.5, "NewYork": 1.4,
               "ERCOT": 0.0, "SPP": 0.0}
# Canada storage: only one grid-scale pumped-hydro plant operates — Sir Adam
# Beck PGS (Niagara, ON, 0.174 GW); Marmora/Meaford/Canyon Creek are proposed,
# not built. Treated like the US PHS fleet (fixed existing + growth-capped max).
# All other Canadian grid storage today is batteries -> Li-Ion, taken from the
# CER EF2026 "Battery Storage" series (canada_power_ef2026.xlsx, CANADA_SCENARIO).
CANADA_PHS_GW_2025 = 0.174

# Coal split: lignite is mine-mouth only (no interregional market). 2025 GW
# per region from plant-level research (ND fleet ~3.9 in MISO, TX ~4.4 in
# ERCOT incl. San Miguel, Red Hills 0.44 in SERC). Remainder of the US Pools
# "Coal" category is hardcoal. US coal capacity now follows the US Coal
# Trajectory Model (Coal_Trajectory_Model.xlsx): each region's 2025 base is
# scaled by the national case path (current shares x national decline) —
# ResidualCapacity = LOW case, TotalAnnualMinCapacity = CENTRAL, MaxCapacity =
# HIGH (the central-low gap is built as boiler refurbishment / life-extension).
# Coal->gas conversions (LOW case) become P_Gas_Steam residual (split by the
# regional coal share). The trajectory is US-only; Canada coal keeps the CER path.
COAL_CATEGORY = "Coal"
LIGNITE_GW_2025 = {"MISO": 3.9, "ERCOT": 4.4, "SERC": 0.44}
SAN_MIGUEL_GW, SAN_MIGUEL_RETIRE_YEAR = 0.41, 2028   # solar+storage conversion
CANADA_LIGNITE_GW = 1.53   # SaskPower life-extension (10/2025), flat to 2040
COAL_TECHS = {"P_Coal_Lignite", "P_Coal_Hardcoal"}

# US nuclear (P_Nuclear) follows the PMK "Total LSR" forecast per region as a
# min-target: residual = 2025 LSR (flat — the fleet does not retire in-horizon),
# TotalAnnualMinCapacity = LSR per year (model builds the announced growth),
# max = LSR x this headroom. Supersedes the US-aggregate Nuclear group-min.
NUCLEAR_MAX_HEADROOM = 1.10

# ---------------------------------------------------------------------------
# Canada: CER "Canada's Energy Future 2026" (canada_power_ef2026.xlsx).
# National primary_fuel capacity trajectory (Current Measures) drives residual
# 2025 + min/max guardrails. The CER has no gas technology split, so the
# "Natural Gas" total is divided by the operating-fleet shares from the GEM
# Global Oil & Gas Plant Tracker (Jan 2026): CC 62 / GT 16 / ST 21.5 /
# IC 0.5 % (gas-fired units; pure-oil units belong to the CER Oil category).
# Solar is split utility/distributed by the CER technology-resolution ratio.
# Canadian hardcoal (NS/NB/AB, ~2.5 GW in 2025) follows the CER decline to 0
# by 2035; SK lignite stays at the SaskPower 1.53 GW flat.
# ---------------------------------------------------------------------------
CANADA_SRC = os.path.join(HERE, "canada_power_ef2026.xlsx")
CANADA_SCENARIO = "Current Measures"
CANADA_GAS_SPLIT = {"P_Gas_CCGT": 0.62, "P_Gas_OCGT": 0.16,
                    "P_Gas_Steam": 0.215, "P_Gas_Engines": 0.005}
CANADA_TECH = {  # CER primary_fuel category -> model tech (gas/solar special)
    "Hydro / Wave / Tidal": "P_Hydro_Reservoir",
    "Uranium": "P_Nuclear",
    "Wind": "P_Wind_Onshore_Avg",
    "Oil": "P_Oil",
    "Biomass / Geothermal": "P_Biomass",
}
CANADA_SOLAR_DISTRIBUTED_SHARE = 0.24   # CER technology res: 1.70/7.15 in 2025
# Doubled funnel vs the US (trajectory deliberately rougher): +/-4% to 2028,
# widening to -20%/+60% at 2035 (hydro +20%), held from there (CER data runs
# through 2040 so no post-2035 extrapolation is needed).
CANADA_MARGIN_NEAR = 0.04
CANADA_MARGIN_2035_MIN = 0.80
CANADA_MARGIN_2035_MAX_DEFAULT = 1.60
CANADA_MARGIN_2035_MAX_PER_TECH = {"P_Hydro_Reservoir": 1.20}

def canada_margins(year, tech=None):
    y = min(year, 2035)
    mx_2035 = CANADA_MARGIN_2035_MAX_PER_TECH.get(tech, CANADA_MARGIN_2035_MAX_DEFAULT)
    if y <= 2028:
        mn, mx = 1 - CANADA_MARGIN_NEAR, 1 + CANADA_MARGIN_NEAR
    else:
        frac = (y - 2028) / (2035 - 2028)
        mn = (1 - CANADA_MARGIN_NEAR) + (CANADA_MARGIN_2035_MIN - (1 - CANADA_MARGIN_NEAR)) * frac
        mx = (1 + CANADA_MARGIN_NEAR) + (mx_2035 - (1 + CANADA_MARGIN_NEAR)) * frac
    if tech == "P_Nuclear":
        mn = 1.0
    return mn, mx

def read_canada_cer():
    """Return ({tech: {year: GW}}, {year: hardcoal_GW}) from the CER workbook."""
    df = pd.read_excel(CANADA_SRC, sheet_name="Capacity")
    df["MW"] = pd.to_numeric(df["MW"], errors="coerce")
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
    sel = df[(df["Scenario"] == CANADA_SCENARIO) & (df["Year"].between(2025, 2040))]
    pf = sel[(sel["Resolution"] == "primary_fuel") & (sel["Region_Code"] == "CA")]

    traj = {}
    def series(variable):
        s = pf[pf["Variable"] == variable]
        return {int(y): float(m) / 1000.0 for y, m in zip(s["Year"], s["MW"])}

    for var, tech in CANADA_TECH.items():
        traj[tech] = series(var)
    gas = series("Natural Gas")
    for tech, share in CANADA_GAS_SPLIT.items():
        traj[tech] = {y: v * share for y, v in gas.items()}
    solar = series("Solar")
    traj["P_PV_Utility_Avg"] = {y: v * (1 - CANADA_SOLAR_DISTRIBUTED_SHARE) for y, v in solar.items()}
    traj["P_PV_Rooftop_Commercial"] = {y: v * CANADA_SOLAR_DISTRIBUTED_SHARE for y, v in solar.items()}

    # Hardcoal: technology resolution, provinces ex-SK (SK = lignite), no CCUS
    tech_res = sel[(sel["Resolution"] == "technology") &
                   (sel["Variable"] == "Coal and Coke") &
                   (~sel["Region_Code"].isin(["CA", "SK"]))]
    hardcoal = tech_res.groupby("Year")["MW"].sum() / 1000.0
    hardcoal = {int(y): float(v) for y, v in hardcoal.items()}
    return traj, hardcoal

def read_canada_storage():
    """Return {year: GW} of CER EF2026 'Battery Storage' for national Canada."""
    df = pd.read_excel(CANADA_SRC, sheet_name="Capacity")
    df["MW"] = pd.to_numeric(df["MW"], errors="coerce")
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
    s = df[(df["Scenario"] == CANADA_SCENARIO) &
           (df["Resolution"] == "primary_fuel") &
           (df["Region_Code"] == "CA") &
           (df["Variable"] == "Battery Storage") &
           (df["Year"].between(2025, 2040))]
    return {int(y): float(m) / 1000.0 for y, m in zip(s["Year"], s["MW"])}

# US nuclear: PMK "Operating and Forecast Through 2040". The Summary "Total LSR"
# = non-SMR nuclear, cumulative by in-service date, US only (AnR + IRP/RFP
# sheets). Reconstructed per unit and mapped to the model's US pool regions
# (validated to reproduce the Summary Total-LSR row exactly). SMR units excluded.
def _nuc_anr_region(iso, nerc, state, subregion=None):
    if subregion and str(subregion).endswith("-CN"):
        return "Canada"  # e.g. MRO-CN (Saskatchewan SMR): filed in the US sheet but Canadian
    if iso == "CAISO":
        return "California" if state == "CA" else "WECC"  # Palo Verde (AZ) sells into CAISO but is WECC
    m = {"PJM": "PJM", "MISO": "MISO", "ERCOT": "ERCOT",
         "New York": "NewYork", "New England": "NewEngland", "SPP": "SPP"}
    if iso in m:
        return m[iso]
    if nerc == "SERC": return "SERC"
    if nerc == "TRE":  return "ERCOT"
    if nerc == "RFC":  return "MISO" if state == "MI" else "PJM"
    if nerc == "WECC": return "California" if state == "CA" else "WECC"
    if nerc == "NPCC": return "NewYork" if state == "NY" else "NewEngland"
    if nerc == "MRO":  return "SPP" if state in ("NE", "KS", "SD", "ND") else "MISO"
    return None

def _nuc_irp_region(area, state):
    p = (area or "").split("_")[0]
    if p == "WECC":
        return "California" if "(CA)" in (state or "") else "WECC"
    return {"SPP": "SPP", "MISO": "MISO", "NY": "NewYork", "isoNE": "NewEngland",
            "PJM": "PJM", "SERC": "SERC", "ERCOT": "ERCOT"}.get(p)

def read_nuclear_lsr():
    """Return {region: {year: GW}} of US Large-Scale-Reactor (non-SMR) nuclear,
    cumulative installed by in-service date, 2025-2040 (US pool regions only)."""
    wb = openpyxl.load_workbook(NUCLEAR_SRC, data_only=True, read_only=True)
    out = {r: {y: 0.0 for y in YEARS} for r in
           ("California", "WECC", "SPP", "MISO", "ERCOT", "SERC", "PJM", "NewYork", "NewEngland")}
    # AnR sheet: C=SMR(2) E=status(4) K=state(10) L=iso(11) M=cap(12) O=date(14) V=nerc(21)
    for row in wb["US AnR Units"].iter_rows(min_row=2, values_only=True):
        smr, state, iso, cap, o, nerc = row[2], row[10], row[11], row[12], row[14], row[21]
        sub = row[22] if len(row) > 22 else None
        if smr == "SMR" or cap is None or not hasattr(o, "year"):
            continue
        reg = _nuc_anr_region(iso, nerc, state, sub)
        if reg is None or reg == "Canada":   # Canadian units come from the CER block
            continue
        for y in YEARS:
            if o.year <= y:
                out[reg][y] += cap / 1000.0   # MW -> GW
    # IRP/RFP sheet: E=status(4) F=cap(5) G=year(6) I=state(8) K=area(10) M=tech(12)
    for row in wb["US IRP_RFP Units"].iter_rows(min_row=2, values_only=True):
        tech, cap, gy, state, area = row[12], row[5], row[6], row[8], row[10]
        if tech == "SMR" or cap is None or gy is None:
            continue
        reg = _nuc_irp_region(area, state)
        if reg is None:
            continue
        for y in YEARS:
            if gy <= y:
                out[reg][y] += cap / 1000.0
    wb.close()
    return out

def read_coal_trajectory():
    """Return national {case: {year: GW}} from Coal_Trajectory_Model.xlsx
    (Trajectory sheet). Coal-fired rows: Low=20, Central=12, High=25;
    Conversion Low=21. Columns C..R = 2025..2040."""
    wb = openpyxl.load_workbook(COAL_SRC, data_only=True, read_only=True)
    ws = wb["Trajectory"]
    def rowvals(r):
        return {YEARS[i]: float(ws.cell(row=r, column=3 + i).value) for i in range(len(YEARS))}
    out = {"coal_low": rowvals(20), "coal_central": rowvals(12),
           "coal_high": rowvals(25), "conv_low": rowvals(21)}
    wb.close()
    return out

def read_nuclear_smr():
    """Return (smr_max, smr_min) per region/year (GW) for P_Nuclear_SMR.
    max = every SMR unit (AnR any status + IRP/RFP), cumulative by in-service
    date — reproduces the Summary 'Total SMR' (~18.7 GW US 2040). min = only the
    firmly-committed AnR 'Development' SMRs (~2.8 GW US 2040); 'Announced' AnR and
    utility-plan IRP SMRs are excluded as not-yet-very-far-ahead."""
    wb = openpyxl.load_workbook(NUCLEAR_SRC, data_only=True, read_only=True)
    regs = ("California", "WECC", "SPP", "MISO", "ERCOT", "SERC", "PJM", "NewYork", "NewEngland", "Canada")
    smr_max = {r: {y: 0.0 for y in YEARS} for r in regs}
    smr_min = {r: {y: 0.0 for y in YEARS} for r in regs}
    for row in wb["US AnR Units"].iter_rows(min_row=2, values_only=True):
        smr, status, state, iso, cap, o, nerc = row[2], row[4], row[10], row[11], row[12], row[14], row[21]
        sub = row[22] if len(row) > 22 else None
        if smr != "SMR" or cap is None or not hasattr(o, "year"):
            continue
        reg = _nuc_anr_region(iso, nerc, state, sub)
        if reg is None:
            continue
        for y in YEARS:
            if o.year <= y:
                smr_max[reg][y] += cap / 1000.0
                if status == "Development":          # firmly-committed subset -> min
                    smr_min[reg][y] += cap / 1000.0
    for row in wb["US IRP_RFP Units"].iter_rows(min_row=2, values_only=True):
        tech, cap, gy, state, area = row[12], row[5], row[6], row[8], row[10]
        if tech != "SMR" or cap is None or gy is None:
            continue
        reg = _nuc_irp_region(area, state)
        if reg is None:
            continue
        for y in YEARS:
            if gy <= y:
                smr_max[reg][y] += cap / 1000.0      # utility IRP SMR -> max only
    wb.close()
    return smr_max, smr_min

# Restool potential column -> all model tech variants in that family.
# Each variant gets a share of the per-region potential (placeholder split until
# the resource-graded breakdown file lands): 30% Opt / 40% Avg / 30% Inf for
# Utility PV and Onshore Wind. The "rep" is the guardrail/funnel carrier and is
# filled FIRST (best-yield _Opt sites); capacity beyond the rep's share-of-potential
# cascades into the "overflow" classes in order: _Avg, then _Inf. This keeps the
# observed/near-term fleet on the high-CF Opt class instead of defaulting to Avg.
# Rooftop maps to a single power tech (P_PV_Rooftop_Commercial) — the area-based
# "A_Rooftop_*" entries are not touched here.
RESTOOL_MAP = {
    "PV Capacity [GW]": {
        "rep": "P_PV_Utility_Opt",
        "overflow": ["P_PV_Utility_Avg", "P_PV_Utility_Inf"],
        "variants": {
            "P_PV_Utility_Opt": 0.30,
            "P_PV_Utility_Avg": 0.40,
            "P_PV_Utility_Inf": 0.30,
        },
    },
    "Wind Capacity [GW]": {
        "rep": "P_Wind_Onshore_Opt",
        "overflow": ["P_Wind_Onshore_Avg", "P_Wind_Onshore_Inf"],
        "variants": {
            "P_Wind_Onshore_Opt": 0.30,
            "P_Wind_Onshore_Avg": 0.40,
            "P_Wind_Onshore_Inf": 0.30,
        },
    },
    "Rooftop Capacity [GW]": {
        "rep": None,
        "variants": {"P_PV_Rooftop_Commercial": 1.00},
    },
}
GUARDRAIL_REP_TO_RESTOOL_COL = {
    info["rep"]: col for col, info in RESTOOL_MAP.items() if info["rep"]
}
# Share that each variant takes of its parent restool column.
TECH_SHARE = {v: s for info in RESTOOL_MAP.values() for v, s in info["variants"].items()}
ALL_RESTOOL_TECHS = sorted(TECH_SHARE.keys())

# Full set of techs this script directly writes positive caps for.
NUCLEAR_SMR_TECHS = {"P_Nuclear_SMR"}
ALL_MANAGED_TECHS = MODEL_TECHS_GUARDRAIL | set(ALL_RESTOOL_TECHS) | COAL_TECHS | HYDRO_TECHS | STORAGE_TECHS | NUCLEAR_SMR_TECHS | {"P_Gas_CCGT_Residual", "P_SOFC"}

# Techs whose caps are owned by other scripts — leave them alone here.
EXTERNAL_OWNERS = {
    # add_offshore_wind_bounds.py
    "P_Wind_Offshore_Deep", "P_Wind_Offshore_Shallow", "P_Wind_Offshore_Transitional",
    # add_egs.py
    "P_EGS_R1", "P_EGS_R2", "P_EGS_R3", "P_EGS_R4",
    # add_biomass_fleet.py (US pools: historic residual, no growth; Canada via CER block)
    "P_Biomass",
}

# Every remaining power-producing tech (P_*, CHP_*) gets MaxCap=0 in the NA
# pool regions + Canada. Storage (D_*, S_*), sector-coupling and demand techs
# stay untouched. Rooftop Residential explicitly zeroed (only Commercial is
# allowed to carry the rooftop potential).
TECH_SHEET = lambda: os.path.join(DATA_REPO, "Data", "Parameters", "00_Sets&Tags", "Sets_Technology.csv")

# Retirement: LINEAR, as share of the 2025 base per year (5% of the 2025
# fleet retires each year — constant absolute retirement, not geometric;
# geometric 0.95^y made retirements slow down over time).
RETIRE_RATE_DEFAULT = 0.05
RETIRE_RATE_PER_TECH = {
    "P_Nuclear": 0.0,
    # Young fleet (mostly 2018+, 30-35 yr life): real retirements start ~2050;
    # 1%/yr covers attrition/failures. 5% forced phantom replacement capex.
    "P_PV_Utility_Opt": 0.01,
    # 2008-2012 build wave hits 20-25 yr in 2030-2037, partly repowered:
    # 45% of the 2025 fleet gone by 2040.
    "P_Wind_Onshore_Opt": 0.03,
    # 15-year battery service life: the whole 2025 fleet retires by 2040.
    "D_Battery_Li-Ion": 1.0 / 15.0,
}

def residual_factor(tech, year):
    rate = RETIRE_RATE_PER_TECH.get(tech, RETIRE_RATE_DEFAULT)
    return max(0.0, 1.0 - rate * (year - 2025))

# US gas retirements, CCGT + OCGT combined (announced pipeline, national GW/yr).
# The US CCGT fleet is young (1999-2005 build wave), so pre-2035 retirements are
# low; post-2035 the schedule ramps toward ~13 GW/yr as that wave hits end of
# life. Allocated proportional to each pool's 2025 CCGT+SCCT fleet (= uniform
# survival fraction on the national fleet). Steam is NOT on this schedule: its
# residual follows the capacity file's own per-region ST decline (net ~= gross
# for steam — no new gas-steam is built), extended post-2035 by the 2030-2035
# regional trend. Engines keep the default rate (tiny, net-growing fleet).
# Canada stays on the default rate (CER data, not this US schedule).
GAS_RETIREMENT_GW = {2026: 0.8, 2027: 0.8, 2028: 1.6, 2029: 2.7, 2030: 2.5,
                     2031: 2.2, 2032: 3.3, 2033: 3.5, 2034: 4.7, 2035: 5.5,
                     2036: 11.0, 2037: 12.0, 2038: 13.0, 2039: 13.5, 2040: 14.0}
GAS_SCHEDULE_TECHS = {"P_Gas_CCGT", "P_Gas_OCGT"}
GAS_SCHEDULE_CATS = ("Natural Gas - CCCT", "Natural Gas - SCCT")

# ---------------------------------------------------------------------------
# Demand-consistency scaling of the guardrail funnel. The US Pools capacity
# trajectories were built for the US Pools generation forecast; our model demand
# comes from the FEL file instead. Per region+year the funnel BASIS is scaled by
#   ratio = FEL busbar demand [TWh] / US-Pools generation [TWh]  (Storage output
# excluded from the generation sum — discharge would double-count),
# so if FEL demand is 10% below the Pools generation, the funnel basis is the
# capacity file value x 0.90. The funnel MIN is then a constant -2% below that
# scaled basis through 2035 (NO widening to -10% any more); from 2036 the min
# widens gradually from the 2035 value to -10% below it at 2040. The max keeps
# its existing widening (and the CA/NE/NY exact-file rule), on the scaled basis.
# Residuals are NOT scaled (the existing fleet is physical) — a residual floor
# on the max keeps min<=max feasible everywhere.
FEL_DEMAND_XLSX = os.path.join(HERE, "base_fel_v260707_v2.xlsx")
if "--fel" in sys.argv:
    FEL_DEMAND_XLSX = os.path.join(HERE, sys.argv[sys.argv.index("--fel") + 1])
FEL_REGION_MAP = {   # FEL geo_code -> US pool (FRCC folds into SERC; Canada not scaled)
    "US_R_CALIFORNIA": "California", "US_R_ERCOT": "ERCOT", "US_R_ISONE": "NewEngland",
    "US_R_MISO": "MISO", "US_R_NYISO": "NewYork", "US_R_PJM": "PJM",
    "US_R_SERC": "SERC", "US_R_SPP": "SPP", "US_R_WECC": "WECC", "US_R_FRCC": "SERC",
}
FUNNEL_MIN_MARGIN = 0.02      # constant -2% below the demand-scaled basis (<= 2035)
GAS_MIN_BLEND_FLOOR = 0.93
if GAS_MIN_BLEND_FLOOR_OVERRIDE is not None:
    GAS_MIN_BLEND_FLOOR = GAS_MIN_BLEND_FLOOR_OVERRIDE    # gas blend floor: the demand-ratio discount on the gas
                              # min never pushes the multiplier below 0.93 x file
                              # (halves the max discount; removes the 2033-35
                              # additions dip the full ratio phase-in produced)
GAS_PEG_MARGIN = 0.02         # gas funnel pegged to the capacities file +/-2% through
                              # GAS_MIN_PIN_UNTIL: min = file x0.98, max = file x1.02
                              # (no demand scaling / widening for gas before 2031)
# GasPlants/USA annual-addition cap (turbine supply constraint): near-term
# turbine deliveries are tight (6/9/10 GW 2026-2028, 35 GW 2029-2030), then the
# supply chain expands to 60 GW/yr by 2035. 2025 stays open (start year).
GAS_GROUP_CAP_GW = {2025: 65.0, 2026: 6.0, 2027: 9.0, 2028: 10.0, 2029: 35.0,
                    2030: 35.0, 2031: 47.0, 2032: 51.0, 2033: 55.0, 2034: 60.0}
GAS_GROUP_CAP_FROM_2035 = 65.0
GAS_CLAMP_BUFFER_GW = 2.0     # forced new-build (min - residual increments) of the
                              # CCGT/OCGT/Steam group must stay this far inside the
                              # year's group addition cap
GAS_MIN_PIN_UNTIL = 2030      # gas min = pegged capacity-file value through this year
                              # (announced/under-construction pipeline), then blends
                              # linearly into the demand-scaled funnel by 2035.
                              # Post-2035 every min HOLDS its 2035 level (no decline);
                              # gas additionally keeps growing (see GAS_MIN_TREND_FRACS).
# Post-2035 gas floor growth: fraction of each (region,tech)'s own 2030-2035 min
# slope added per year. 0.5 at 2036 declining linearly to 0.278 at 2040 = ~9 down
# to ~5 GW/yr nationally (the full national slope 2030-35 is ~18 GW/yr).
GAS_MIN_TREND_FRACS = {2036: 0.500, 2037: 0.444, 2038: 0.389, 2039: 0.333, 2040: 0.278}


def read_fel_demand():
    """Return {region: {year: busbar_TWh}} from the FEL demand workbook
    (_Data_Grid, all sectors summed, US pools only)."""
    df = pd.read_excel(FEL_DEMAND_XLSX, "_Data_Grid")
    df = df[df["geo_code"].isin(FEL_REGION_MAP)]
    df["Region"] = df["geo_code"].map(FEL_REGION_MAP)
    g = df.groupby(["Region", "year"])["busbar_twh"].sum()
    out = {}
    for (r, y), v in g.items():
        out.setdefault(r, {})[int(y)] = float(v)
    return out

def hydro_sheet_traj(values_2025_2035):
    """Hydro capacity trajectory: the US Pools sheet 2025-2035 verbatim, with
    2036-2040 linearly extrapolated from the 2030-2035 trend. `values_2025_2035`
    is a {year: GW} dict that must cover 2025..2035."""
    s = dict(values_2025_2035)
    slope = (s[2035] - s.get(2030, s[2035])) / 5.0
    for y in range(2036, 2041):
        s[y] = max(0.0, s[2035] + slope * (y - 2035))
    return s

# Gas technologies: TotalAnnualMaxCapacity must never decrease over the years
# — a dipping cap forces the model to scrap capacity it was allowed (or
# needed) to build earlier, which is infeasible with NewCapacity >= 0.
MONOTONIC_MAX_TECHS = {"P_Gas_CCGT", "P_Gas_OCGT", "P_Gas_Steam", "P_Gas_Engines"}

# --- Colleague feedback (2026-06) -----------------------------------------
# (1) California / New York / New England get NO gas additions beyond the US
#     capacities file: pin the gas TotalAnnualMaxCapacity to the sheet value (max
#     widening factor -> 1.0, no post-2035 growth). The declining-fleet residual +
#     the monotonic floor still apply, so the max = the file's own peak held flat
#     (feasible; just no funnel-driven new gas).
GAS_NO_ADD_REGIONS = {"California", "NewYork", "NewEngland"}
GAS_TECHS = {"P_Gas_CCGT", "P_Gas_OCGT", "P_Gas_Steam", "P_Gas_Engines"}
# (2) SERC has a weak SE onshore-wind resource and builds little in reality; the
#     restool potential (~56 GW) overstates it. Cap the SERC onshore potential to
#     narrow the whole upper funnel (rep + overflow classes scale with it). Floor
#     is set by the existing fleet sitting on _Opt (residual ~3.8 GW, Opt share
#     0.30 -> needs pot >= ~12.7), so 15 GW is about as tight as stays feasible.
ONSHORE_POT_OVERRIDE_GW = {"SERC": 15.0}

YEARS = list(range(2025, 2041))
DATE, WHO = "2026-06-04", "Konstantin Loffler <kl@wip.tu-berlin.de>"

# Max-side widening cap at 2035. Default is 1.30 (+30%). Hydro is narrower —
# reservoirs are physically constrained, the funnel should not pretend they can
# triple in capacity.
MAX_WIDEN_2035_DEFAULT = 1.30
MAX_WIDEN_2035_PER_TECH = {"P_Hydro_Reservoir": 1.10}

# Restool upside (_Opt/_Inf) ramp. The variants open linearly from 0 at
# RESTOOL_RAMP_START to full share-of-potential at RESTOOL_RAMP_END. With
# switch_investLimit on, SC2 (+NewRESCapacity) paces the per-year build, so the
# funnel can be widened without re-introducing the 2036 spike: open from 2031 and
# reach full potential by 2045 (was 2035->2050). The rep (_Avg) max stays pinned
# through 2035, then extends its 2030-2035 trend.
RESTOOL_RAMP_START = 2031
RESTOOL_RAMP_END = 2045
def restool_frac(y):
    return max(0.0, min(1.0, (y - RESTOOL_RAMP_START) / (RESTOOL_RAMP_END - RESTOOL_RAMP_START)))

# Canada gets a much slower upside ramp: its restool classes are NOT tied to a
# guardrail rep, so the ramp slope IS the annual-addition ceiling. 2032->2070
# = ~5 GW/yr (PV Opt) / ~3 GW/yr (wind Opt) - and because the ceiling stays
# binding through 2040 the buildout cannot saturate-and-cliff mid-horizon.
CANADA_RESTOOL_RAMP_START = 2032
CANADA_RESTOOL_RAMP_END = 2070
CANADA_RESTOOL_SLOPE_GW = 4.0   # hard ceiling-growth cap, GW/yr per class: Canada's
                                # PV potential is so large (~1450 GW) that a time
                                # window alone still allowed ~11 GW/yr
def canada_restool_frac(y):
    return max(0.0, min(1.0, (y - CANADA_RESTOOL_RAMP_START) /
                        (CANADA_RESTOOL_RAMP_END - CANADA_RESTOOL_RAMP_START)))


# Annual MAX growth applied after 2035 for techs WITHOUT a restool potential
# target. PV/Wind interpolate to the restool potential at 2040 (different code
# path) — these rates only apply to thermal + hydro. Without this growth the
# 2036-2040 cap would be flat which contradicts e.g. the Nuclear/USA target
# floor reaching 129.8 GW in 2040.
POST_2035_MAX_GROWTH = {
    "P_Gas_CCGT":         0.05,
    "P_Gas_OCGT":         0.05,   # flexible gas must keep room to grow post-2035
    "P_Gas_Engines":      0.05,
    "P_Nuclear":          0.05,
    "P_Hydro_Reservoir":  0.02,
}

# Nuclear/USA target floor (GroupTotalAnnualMinCapacity).
NUCLEAR_USA_GROUP_MIN = {
    2035: 116.0,   2036: 118.3,   2037: 121.43,
    2038: 123.932, 2039: 123.932, 2040: 129.832,
}


def margins(year, tech=None, region=None):
    """(min_factor, max_factor) widening from +/-2% (<=2028) to -10%/+max(2035)
    where max(2035) is per-tech (1.30 default, 1.10 for hydro). Year capped at
    2035 here; post-2035 growth is handled separately by POST_2035_MAX_GROWTH.
    Gas in GAS_NO_ADD_REGIONS gets max_factor pinned to 1.0 (no additions beyond
    the US capacities file)."""
    y = min(year, 2035)
    mx_2035 = MAX_WIDEN_2035_PER_TECH.get(tech, MAX_WIDEN_2035_DEFAULT)
    if y <= 2028:
        mn, mx = 0.98, 1.02
    else:
        frac = (y - 2028) / (2035 - 2028)
        mn = 0.98 + (0.90 - 0.98) * frac
        mx = 1.02 + (mx_2035 - 1.02) * frac
    if tech == "P_Nuclear":
        mn = 1.0
    if region in GAS_NO_ADD_REGIONS and tech in GAS_TECHS:
        mx = 1.0   # cap gas at the sheet value: no endogenous additions
    return mn, mx


def read_restool_potentials():
    """Return {region: {col: GW}} from both NA + Canada potential CSVs."""
    out = {}
    for path in (POT_NA, POT_CA):
        df = pd.read_csv(path)
        for _, row in df.iterrows():
            region = str(row["region"])
            out[region] = {col: float(row[col]) for col in row.index
                           if col != "region" and not pd.isna(row[col])}
    return out


def main():
    df = pd.read_excel(SRC)
    # 2026-06 file update: category column renamed to "Fuel + Technology",
    # measure column is unnamed (3rd position).
    df = df.rename(columns={df.columns[1]: "Category", df.columns[2]: "Measure"})
    cap = df[df["Measure"] == "Capacity MW"].copy()
    pot = read_restool_potentials()
    # Colleague feedback: narrow the SERC onshore upper funnel by capping its
    # restool wind potential (rep share-cap + overflow targets scale with it).
    for _reg, _cap_gw in ONSHORE_POT_OVERRIDE_GW.items():
        if _reg in pot and "Wind Capacity [GW]" in pot[_reg]:
            pot[_reg]["Wind Capacity [GW]"] = min(pot[_reg]["Wind Capacity [GW]"], _cap_gw)

    def gw(region, fuel, year):
        sel = cap[(cap["Pool-Regions"] == region) & (cap["Category"] == fuel)]
        if sel.empty or year not in sel.columns:
            return 0.0
        return max(0.0, float(sel.iloc[0][year])) / 1000.0   # MW->GW, clamp >=0

    pool_regions = sorted(cap["Pool-Regions"].unique())
    extra_regions = [r for r in pot.keys() if r not in pool_regions]  # Canada etc.

    res_rows, min_rows, max_rows = [], [], []

    # US coal trajectory (national) + per-region 2025 coal shares. Each region's
    # coal base is scaled by the national case ratio (1.0 at 2025 since the
    # uncertainty band is 0 there). Coal->gas conversions (LOW case) are split by
    # the regional coal share and added to P_Gas_Steam residual.
    coal_traj = read_coal_trajectory()
    c0 = coal_traj["coal_low"][2025]   # = central[2025] = high[2025] = 171
    coal_ratio_low     = {y: coal_traj["coal_low"][y] / c0     for y in YEARS}
    coal_ratio_central = {y: coal_traj["coal_central"][y] / c0 for y in YEARS}
    coal_ratio_high    = {y: coal_traj["coal_high"][y] / c0    for y in YEARS}
    conv_low = coal_traj["conv_low"]
    # The Natural-Gas-ST base already bakes in the coal->gas conversions that had
    # happened by the base year (conv_low[2025] ~ 11 GW), so only the *additional*
    # conversions since 2025 may be added to the P_Gas_Steam residual -- adding the
    # absolute conv_low[y] would double-count the base-year fleet. Floor at 0 so a
    # net decline (converted plants retiring) is handled by the base residual decay,
    # not by subtracting capacity here.
    conv_base = conv_low.get(2025, 0.0)
    conv_add = {y: max(0.0, conv_low.get(y, 0.0) - conv_base) for y in YEARS}
    coal_2025 = {r: gw(r, COAL_CATEGORY, 2025) for r in pool_regions}
    coal_tot = sum(coal_2025.values())
    coal_share = {r: (coal_2025[r] / coal_tot if coal_tot else 0.0) for r in pool_regions}

    # CCGT+OCGT survival fraction from the national retirement schedule:
    # proportional allocation over the 2025 fleet = one national fraction path.
    nat_gas_2025 = sum(gw(r, c, 2025) for r in pool_regions for c in GAS_SCHEDULE_CATS)
    gas_surv, _cum = {2025: 1.0}, 0.0
    for y in YEARS[1:]:
        _cum += GAS_RETIREMENT_GW.get(y, 0.0)
        gas_surv[y] = max(0.0, 1.0 - _cum / nat_gas_2025) if nat_gas_2025 else 1.0

    # Demand-consistency ratio per (region, year): FEL busbar demand / US-Pools
    # generation (Output MWh, Storage excluded). Missing region/year -> 1.0.
    gen_rows = df[df["Measure"] == "Output MWh"]
    gen_rows = gen_rows[gen_rows["Category"] != STORAGE_CATEGORY]
    fel = read_fel_demand()
    fel_gen_ratio = {}
    print("Funnel demand/generation scaling (FEL busbar TWh / Pools generation TWh):")
    for r in pool_regions:
        gsum = gen_rows[gen_rows["Pool-Regions"] == r]
        ratios = {}
        for y in range(2025, 2036):
            gen_twh = max(0.0, float(gsum[y].sum())) / 1e6 if y in gsum.columns else 0.0
            dem_twh = fel.get(r, {}).get(y, 0.0)
            # RAW ratio stored; capped at the use sites: the MIN side always caps
            # at 1.0 (importers' deficit is met by endogenous trade), the MAX side
            # caps at 1.0 unless --max-upscale (dc_high) lets extra demand raise it.
            ratios[y] = dem_twh / gen_twh if (gen_twh > 0 and dem_twh > 0) else 1.0
        fel_gen_ratio[r] = ratios
        print(f"  {r:12} 2025: {ratios[2025]:.3f}  2030: {ratios[2030]:.3f}  2035: {ratios[2035]:.3f}")

    # US nuclear LSR trajectory per pool region (Canada nuclear = CER block).
    nuclear_lsr = read_nuclear_lsr()
    # SMR (P_Nuclear_SMR): full forecast (max) + committed-Development subset (min).
    nuclear_smr_max, nuclear_smr_min = read_nuclear_smr()

    # 1) Guardrail funnel (5 fuel groups x US pool regions). We also accumulate
    #    per (region, year) the "excess" between the rep guardrail trajectory
    #    (max and min) and the rep's share of the per-region restool potential.
    #    Max excess spills into _Opt (up to Opt's headroom). Min excess flows
    #    through _Opt then _Inf so the per-region min requirement stays
    #    satisfiable when the rep alone cannot carry it.
    excess_by_region_rep_year     = {}   # {(region, rep, year): max_excess_GW}
    excess_min_by_region_rep_year = {}   # {(region, rep, year): min_excess_GW}
    for region in pool_regions:
        # Rep's share of total potential (per family). Used to bound rep_cap
        # at the share and to compute excess (= rep guardrail - share).
        rep_share_pot = {}
        for rep, col in GUARDRAIL_REP_TO_RESTOOL_COL.items():
            tot = pot.get(region, {}).get(col)
            rep_share_pot[rep] = (tot * TECH_SHARE.get(rep, 0.0)) if tot is not None else None
        for fuel, tech in TECH.items():
            if tech == "P_Nuclear":
                continue   # handled in the dedicated LSR block (1b2) below
            base = gw(region, fuel, 2025)
            # Steam residual = the file's own regional ST path (existing fleet;
            # capped at the 2025 base so file-side growth never inflates it),
            # 2036-2040 extended by the 2030-2035 trend via hydro_sheet_traj.
            st_traj = hydro_sheet_traj({y: gw(region, fuel, y) for y in range(2025, 2036)}) \
                if tech == "P_Gas_Steam" else None
            def _res_raw(y):
                if tech in GAS_SCHEDULE_TECHS:
                    return base * gas_surv[y]                      # national schedule
                if tech == "P_Gas_Steam":
                    return min(base, st_traj[y]) + conv_add[y] * coal_share.get(region, 0.0)
                return base * residual_factor(tech, y)
            # Residual paths must be MONOTONE NON-INCREASING for gas: the stepped
            # coal->gas conversion trajectory (conv_add) otherwise creates upward
            # jumps in the existing Steam fleet (e.g. US +5.8 GW in 2030) that the
            # model gets for free - clamp them away so conversions show up as
            # buildable NewCapacity instead.
            _res_path, _prev = {}, None
            for y in YEARS:
                v = _res_raw(y)
                if tech in GAS_TECHS and _prev is not None:
                    v = min(v, _prev)
                _res_path[y] = v; _prev = v
            _res = lambda y: _res_path[y]
            for y in YEARS:
                res_val = _res(y)
                if tech == "P_Gas_CCGT":
                    # existing CCGT fleet -> permit-limited P_Gas_CCGT_Residual (av 0.5);
                    # P_Gas_CCGT carries no residual (it is the new-build headroom).
                    res_rows.append((region, "P_Gas_CCGT_Residual", y, round(res_val, 6)))
                    res_rows.append((region, "P_Gas_CCGT", y, 0.0))
                else:
                    res_rows.append((region, tech, y, round(res_val, 6)))

            # Demand-scaled funnel basis: file value x FEL/Pools ratio (see the
            # FEL_DEMAND_XLSX block). Ratio is defined 2025-2035 (Pools horizon).
            _ratios = fel_gen_ratio.get(region, {})
            def _rmin(y):                       # min side: never above 1.0
                return min(1.0, _ratios.get(min(y, 2035), 1.0))
            def _rmax(y):                       # max side: >1 only with --max-upscale
                r = _ratios.get(min(y, 2035), 1.0)
                if not MAX_UPSCALE:
                    return min(1.0, r)
                if tech in RES_HALF_UPSCALE and r > 1.0:
                    return 1.0 + (r - 1.0) * 0.5   # half-strength RES upscaling
                return r
            def sval(y):                        # demand-scaled basis for the MAX side
                yy = min(y, 2035)
                return gw(region, fuel, yy) * _rmax(y)
            def svalmin(y):                     # demand-scaled basis for the MIN side
                yy = min(y, 2035)
                return gw(region, fuel, yy) * _rmin(y)

            # 2035 funnel anchors for post-2035 interp + min widening
            val_2035 = sval(2035)
            _, mx_at_2035 = margins(2035, tech, region)
            max_2035 = val_2035 * mx_at_2035 *                 (1.0 if (tech in GAS_TECHS and region in GAS_NO_ADD_REGIONS) else econ_max_f(2035))
            # min anchor at 2035: gas gets the same blend floor as the year loop
            # (else the post-2035 trend starts from the unfloored value and the
            # floor DROPS at 2036)
            if tech in GAS_TECHS:
                min_2035 = gw(region, fuel, 2035) * max(GAS_MIN_BLEND_FLOOR,
                           _rmin(2035) * (1.0 - FUNNEL_MIN_MARGIN)) * econ_min_f(2035) * gas_min_relax_f(2035)
            else:
                min_2035 = svalmin(2035) * (1.0 - FUNNEL_MIN_MARGIN) * econ_min_f(2035)
            # slope of the rep max over 2030-2035, extended post-2035 (see below)
            val_2030 = sval(2030)
            _, mx_at_2030 = margins(2030, tech, region)
            avg_slope_post2035 = (max_2035 - val_2030 * mx_at_2030) / 5.0

            target_2040 = rep_share_pot.get(tech)  # rep share x total pot (None if N/A)

            running_max = 0.0   # monotonic floor for MONOTONIC_MAX_TECHS
            running_req_new = 0.0   # persistence: new build already forced by past funnel
                                    # mins above the residual (new capacity lives 30+ yrs,
                                    # so every later max must still accommodate it)
            for y in range(2026, 2041):
                if y <= 2035:
                    _, mx = margins(y, tech, region)
                    if tech in GAS_TECHS and y <= GAS_MIN_PIN_UNTIL:
                        # near-term gas floor pegged to the capacity-file value
                        # -2% (no demand scaling): the 2026-2029 gas pipeline is
                        # announced/under construction, not a band.
                        raw_min = gw(region, fuel, y) * (1.0 - GAS_PEG_MARGIN)
                    elif tech in GAS_TECHS:
                        # blend linearly from the exact-file pin (at GAS_MIN_PIN_UNTIL)
                        # into the demand-scaled -2% funnel (at 2035)
                        frac = (y - GAS_MIN_PIN_UNTIL) / (2035 - GAS_MIN_PIN_UNTIL)
                        blendf = (1.0 - frac) * (1.0 - GAS_PEG_MARGIN) + frac * _rmin(y) * (1.0 - FUNNEL_MIN_MARGIN)
                        raw_min = gw(region, fuel, y) * max(GAS_MIN_BLEND_FLOOR, blendf) * econ_min_f(y) * gas_min_relax_f(y)
                    else:
                        # min: constant -2% below the demand-scaled basis (no widening)
                        raw_min = svalmin(y) * (1.0 - FUNNEL_MIN_MARGIN) * econ_min_f(y)
                    # economic funnel widens the max too - except the no-add gas
                    # regions, whose exact-file cap is a permitting constraint
                    _emax = 1.0 if (tech in GAS_TECHS and region in GAS_NO_ADD_REGIONS) else econ_max_f(y)
                    raw_max = sval(y) * mx * _emax
                    if tech in GAS_TECHS:
                        # gas max pegged to the file +2% through 2030 (no upward
                        # funnel), then blends into the normal (widened) funnel,
                        # reaching it at 2035.
                        pegged = gw(region, fuel, y) * (1.0 + GAS_PEG_MARGIN)
                        if y <= GAS_MIN_PIN_UNTIL:
                            raw_max = pegged
                        else:
                            gfrac = (y - GAS_MIN_PIN_UNTIL) / (2035.0 - GAS_MIN_PIN_UNTIL)
                            raw_max = (1.0 - gfrac) * pegged + gfrac * raw_max
                else:
                    # min: HOLD the 2035 floor flat post-2035 (a declining floor let
                    # gross additions collapse to ~0 in 2036 - retiring capacity
                    # must be replaced, matching real market behaviour). GAS keeps
                    # growing: extend the tech's own 2030-2035 min trend at half
                    # slope, tapering to ~5 GW/yr nationally by 2040.
                    raw_min = min_2035
                    if tech in GAS_TECHS:
                        slope = max(0.0, (min_2035 - gw(region, fuel, 2030)) / 5.0)
                        raw_min += slope * sum(f for k, f in GAS_MIN_TREND_FRACS.items() if k <= y)
                    raw_min *= post35_min_f(y)   # grid ramp keeps widening post-2035
                    if tech in GAS_TECHS and GAS_MIN_RELAX_2040 is not None:
                        # continue the gentle gas relax past its 2035 anchor value
                        raw_min *= gas_min_relax_f(y) / gas_min_relax_f(2035)
                    # max: interp 2035 funnel -> 2040 share-of-potential if a
                    # restool target is available; otherwise compound growth
                    # using POST_2035_MAX_GROWTH (thermal/hydro) or hold flat.
                    if target_2040 is not None and target_2040 > 0:
                        # extend the 2030-2035 max trend post-2035 instead of
                        # ballooning to the full share-of-potential by 2040
                        # (removes the 2035->2036 funnel cliff). Still capped at
                        # the rep share below.
                        raw_max = max_2035 + avg_slope_post2035 * (y - 2035)
                        raw_max *= post35_max_f(y)
                    else:
                        rate = POST_2035_MAX_GROWTH.get(tech, 0.0)
                        if region in GAS_NO_ADD_REGIONS and tech in GAS_TECHS:
                            rate = 0.0   # no post-2035 gas growth in these regions
                        raw_max = max_2035 * (1.0 + rate) ** (y - 2035)
                        if not (tech in GAS_TECHS and region in GAS_NO_ADD_REGIONS):
                            raw_max *= post35_max_f(y)   # no-add gas stays permit-pinned

                # Cap rep tech at its share of total potential. Excess spills
                # into _Opt (and, for min only, then into _Inf) so the (min ≤
                # max) invariant survives capping by the regional potential.
                share_cap = rep_share_pot.get(tech)
                if share_cap is not None:
                    rep_max = min(raw_max, share_cap)
                    rep_min = min(raw_min, share_cap)
                    excess_by_region_rep_year[(region, tech, y)]     = max(0.0, raw_max - share_cap)
                    excess_min_by_region_rep_year[(region, tech, y)] = max(0.0, raw_min - share_cap)
                else:
                    rep_max = raw_max
                    rep_min = raw_min
                # Residuals are exogenous (the fleet cannot retire early), and they
                # are NOT demand-scaled — floor every funnel max at the residual so
                # a scaled-down basis can never force max < installed fleet. The
                # persistence floor adds the new build that PAST funnel mins forced
                # above the residual (it cannot vanish later: 30+ yr lifetime) —
                # without it a min-year followed by a lower-max year is infeasible
                # (e.g. NY CCGT: min forces +0.148 GW new in 2032, the declining
                # exact-file max then sat below residual+0.148 from 2036 on).
                running_req_new = max(running_req_new, rep_min - _res(y))
                rep_max = max(rep_max, _res(y) + max(0.0, running_req_new))
                rep_max_exact = rep_max   # funnel max BEFORE the monotonic peak-hold (= the
                                          # exact per-year (scaled) file value for CA/NE/NY, mx=1.0)
                if tech in MONOTONIC_MAX_TECHS:
                    rep_max = max(rep_max, running_max)
                    running_max = rep_max
                if tech == "P_Gas_CCGT":
                    # Split the CCGT funnel band: existing fleet -> pinned, non-buildable
                    # _Residual (= the decaying residual, av 0.5); P_Gas_CCGT (av 0.8) = the
                    # band MINUS the existing fleet (min = funnel - residual). For the MAX:
                    # CA/NE/NY get EXACTLY the data entry (mx=1.0, no monotonic hold) minus the
                    # fleet -> total CCGT = the file value exactly; other regions use the
                    # widened (monotonic) funnel max minus the fleet -> allowed to build more.
                    rv = _res(y)   # existing fleet on the CCGT+OCGT retirement schedule
                    # _Residual is pinned to the existing fleet: write both max AND min =
                    # ResidualCapacity. Redundant in the model (the residual already floors
                    # it) but makes run-to-run capacity diffs easier to read.
                    max_rows.append((region, "P_Gas_CCGT_Residual", y, round(rv, 6)))
                    min_rows.append((region, "P_Gas_CCGT_Residual", y, round(rv, 6)))
                    cap_max = rep_max_exact if region in GAS_NO_ADD_REGIONS else rep_max
                    new_min = max(0.0, rep_min - rv)
                    new_max = max(FORBID_EPS, cap_max - rv)
                    if new_min > 0:
                        min_rows.append((region, "P_Gas_CCGT", y, round(new_min, 6)))
                    max_rows.append((region, "P_Gas_CCGT", y, round(new_max, 6)))
                else:
                    min_rows.append((region, tech, y, round(rep_min, 6)))
                    max_rows.append((region, tech, y, round(rep_max, 6)))

        # 1b) Coal: split the US Pools "Coal" capacity into lignite (fixed
        #     regional fleet, mine-mouth fuel) and hardcoal (remainder), then
        #     scale each region's 2025 base by the US Coal Trajectory Model case
        #     path: ResidualCapacity = LOW, TotalAnnualMinCapacity = CENTRAL,
        #     MaxCapacity = HIGH (the central-low gap is built as refurbishment /
        #     life-extension). San Miguel (ERCOT lignite, 0.41 GW) converts to
        #     solar+storage. Conversions feed P_Gas_Steam residual (gas block).
        coal_base = gw(region, COAL_CATEGORY, 2025)
        lign_base = min(LIGNITE_GW_2025.get(region, 0.0), coal_base)
        hard_base = max(0.0, coal_base - lign_base)
        for tech, base in (("P_Coal_Lignite", lign_base), ("P_Coal_Hardcoal", hard_base)):
            for y in YEARS:
                eff_base = base
                if tech == "P_Coal_Lignite" and region == "ERCOT" and y >= SAN_MIGUEL_RETIRE_YEAR:
                    eff_base = max(0.0, base - SAN_MIGUEL_GW)
                res_rows.append((region, tech, y, round(eff_base * coal_ratio_low[y], 6)))
                # min = LOW case = the residual path (follow the input file; the
                # CENTRAL-case min forced refurbishment builds from 2031 on)
                min_rows.append((region, tech, y, round(eff_base * coal_ratio_low[y], 6)))
                # FORBID_EPS floor: lignite-free / coal-free regions (base 0) must
                # stay forbidden, else max==0 -> 999999 (FossilPower rule) lets them
                # build UNLIMITED coal/lignite.
                # grid funnel: NO coal refurbishment headroom (max = the LOW
                # residual path). Coal bounds are otherwise style-independent,
                # but the extra links make coal-region exports profitable and
                # the HIGH-case headroom turned into +13 GW coal in grid_high.
                coal_max_ratio = coal_ratio_low[y] if FUNNEL_STYLE == "grid" else coal_ratio_high[y]
                max_rows.append((region, tech, y, round(max(FORBID_EPS, eff_base * coal_max_ratio), 6)))

        # 1b2) Nuclear: PMK "Total LSR" forecast per region (min-target).
        #     ResidualCapacity = 2025 LSR (flat; fleet does not retire in-horizon),
        #     TotalAnnualMinCapacity = LSR per year (model builds the announced
        #     growth, paying nuclear capex), max = LSR x NUCLEAR_MAX_HEADROOM.
        lsr = nuclear_lsr.get(region, {})
        base_lsr = lsr.get(2025, 0.0)
        for y in YEARS:
            tgt = lsr.get(y, base_lsr)
            res_rows.append((region, "P_Nuclear", y, round(base_lsr, 6)))
            min_rows.append((region, "P_Nuclear", y, round(tgt, 6)))
            max_rows.append((region, "P_Nuclear", y, round(tgt * NUCLEAR_MAX_HEADROOM, 6)))

        # 1c) Hydro: split into reservoir + run-of-river (per-region RoR share).
        #     Residual is FLAT at the 2025 fleet (hydro does not retire); the
        #     planned growth is forced via TotalAnnualMinCapacity strict to the
        #     US Pools sheet (2030-2035 trend extrapolated to 2040). Max keeps a
        #     small headroom (HYDRO_MAX_HEADROOM) above the strict min.
        h_traj = hydro_sheet_traj({y: gw(region, HYDRO_CATEGORY, y) for y in range(2025, 2036)})
        ror = HYDRO_ROR_SHARE.get(region, HYDRO_ROR_DEFAULT)
        for tech, sh in (("P_Hydro_Reservoir", 1.0 - ror), ("P_Hydro_RoR", ror)):
            for y in YEARS:
                res_rows.append((region, tech, y, round(sh * h_traj[2025], 6)))
                min_rows.append((region, tech, y, round(sh * h_traj[y], 6)))
                max_rows.append((region, tech, y, round(sh * h_traj[y] * HYDRO_MAX_HEADROOM, 6)))

        # 1d) Storage: split US Pools "Storage" into existing pumped hydro + Li-Ion
        #     BESS. PHS existing fleet persists (residual = min = the 2025 fleet);
        #     its max is a growth-rate-capped ceiling (phs * (1+PHS_MAX_GROWTH)^
        #     (y-2025)) — the real limit on new US PHS is buildout rate, not a
        #     hard pin, so the optimiser may add modest PHS but cannot balloon it
        #     (regions with no 2025 PHS stay ~0). The battery share (Storage -
        #     PHS) and all its growth go to D_Battery_Li-Ion: residual = the 2025
        #     battery fleet (retiring at the default rate), TotalAnnualMinCapacity
        #     forces the US Pools storage trajectory (2030-2035 trend -> 2040) via
        #     the standard funnel, and the max is left open above that floor.
        s_traj = hydro_sheet_traj({y: gw(region, STORAGE_CATEGORY, y) for y in range(2025, 2036)})
        phs = min(PHS_GW_2025.get(region, 0.0), s_traj[2025])
        bess_2025 = max(0.0, s_traj[2025] - phs)
        for y in YEARS:
            res_rows.append((region, "D_PHS", y, round(phs, 6)))
            min_rows.append((region, "D_PHS", y, round(phs, 6)))
            max_rows.append((region, "D_PHS", y, round(max(FORBID_EPS, phs * (1.0 + PHS_MAX_GROWTH) ** (y - 2025)), 6)))
            bess = max(0.0, s_traj[y] - phs)
            mn, mx = margins(y)
            # under --max-upscale (dc_high) the storage ceiling grows with the
            # regional demand ratio - the US Pools trajectory reflects BASE
            # demand and starved high-demand regions of batteries (ERCOT).
            stor_up = max(1.0, fel_gen_ratio.get(region, {}).get(min(y, 2035), 1.0)) if MAX_UPSCALE else 1.0
            res_rows.append((region, "D_Battery_Li-Ion", y, round(bess_2025 * residual_factor("D_Battery_Li-Ion", y), 6)))
            liion_min = round(bess * mn * bess_min_relax_f(y), 6)
            min_rows.append((region, "D_Battery_Li-Ion", y, liion_min))
            # funnel max around the US Pools storage trajectory (was open 999999,
            # which let the optimiser front-load ~100 GW of BESS into 2026);
            # --bess-pin: max = the relaxed min (hard low build-out, no upside)
            liion_max = liion_min if BESS_PIN else round(max(FORBID_EPS, bess * mx * stor_up), 6)
            max_rows.append((region, "D_Battery_Li-Ion", y, max(FORBID_EPS, liion_max)))

        # 2) Overflow restool variants per region (the classes after the rep, in
        #    the order given by RESTOOL_MAP["overflow"], e.g. _Avg then _Inf).
        #    Behaviour:
        #    Max:
        #     - 2025-2035: cap = 0 (do not introduce these variants yet)
        #     - 2036-2040: linear ramp from 0 to (share * total_potential) at 2040
        #     - the first overflow class additionally absorbs the rep's MAX excess,
        #       capped at its remaining headroom below its own share*pot.
        #    Min:
        #     - The rep's MIN excess flows to the first overflow class (bounded by
        #       its max value at year y), then any residual to the second (bounded
        #       by its max). So total per-region min is preserved without ever
        #       requiring more than that variant's own max.
        for col, info in RESTOOL_MAP.items():
            pot_val = pot.get(region, {}).get(col, 0.0)
            rep = info["rep"]
            if rep is None:
                # Single variant (Rooftop) — share fraction flat across years
                for variant, share in info["variants"].items():
                    target = pot_val * share
                    for y in YEARS:
                        max_rows.append((region, variant, y, round(target, 6)))
                continue
            # Ordered overflow classes after the rep: capacity beyond the rep's
            # share cascades into overflow[0] (e.g. _Avg) first, then overflow[1]
            # (_Inf). Each absorbs the rep excess up to its own share-of-potential.
            overflow = info.get("overflow", [])
            first_tech = overflow[0] if len(overflow) > 0 else None
            second_tech = overflow[1] if len(overflow) > 1 else None
            first_target = pot_val * info["variants"].get(first_tech, 0.0)
            second_target = pot_val * info["variants"].get(second_tech, 0.0)
            for y in YEARS:
                frac = restool_frac(y)
                first_base = first_target * frac
                second_base = second_target * frac
                # rep MAX excess spills into the first overflow class, capped at its
                # remaining headroom below its share*pot.
                exc_max = excess_by_region_rep_year.get((region, rep, y), 0.0)
                first_max = first_base + min(exc_max, max(0.0, first_target - first_base))
                second_max = second_base
                # rep MIN excess flows first -> second, each bounded by its own max.
                exc_min = excess_min_by_region_rep_year.get((region, rep, y), 0.0)
                first_min = min(exc_min, first_max)
                second_min = min(max(0.0, exc_min - first_min), second_max)
                if first_tech is not None:
                    max_rows.append((region, first_tech, y, round(first_max, 6)))
                    if first_min > 0:
                        min_rows.append((region, first_tech, y, round(first_min, 6)))
                if second_tech is not None:
                    max_rows.append((region, second_tech, y, round(second_max, 6)))
                    if second_min > 0:
                        min_rows.append((region, second_tech, y, round(second_min, 6)))

    # 3) Canada: CER EF2026 trajectory drives residual 2025 + min/max funnel
    #    (doubled margins vs the US). Restool keeps the _Opt/_Inf upside ramps
    #    2036-2040 on top; the rep variants and rooftop are owned by the CER
    #    block, so they are skipped in the restool loop below.
    canada_techs = set()
    if os.path.exists(CANADA_SRC) and "Canada" in extra_regions:
        traj, ca_hardcoal = read_canada_cer()
        canada_techs = set(traj.keys()) | {"P_Hydro_RoR"}   # RoR also CER-owned
        running_min_gas = {}   # per-tech monotone floor for the Canadian gas min
        for tech, ser in sorted(traj.items()):
            base = ser.get(2025, 0.0)
            # Hydro: same treatment as the US pools — flat residual, strict min
            # to the CER trajectory, split reservoir/run-of-river.
            if tech == "P_Hydro_Reservoir":
                ror = HYDRO_ROR_SHARE.get("Canada", HYDRO_ROR_DEFAULT)
                for ht, sh in (("P_Hydro_Reservoir", 1.0 - ror), ("P_Hydro_RoR", ror)):
                    for y in YEARS:
                        val = ser.get(y, base)
                        res_rows.append(("Canada", ht, y, round(sh * base, 6)))
                        min_rows.append(("Canada", ht, y, round(sh * val, 6)))
                        max_rows.append(("Canada", ht, y, round(sh * val * HYDRO_MAX_HEADROOM, 6)))
                continue
            running_max = 0.0
            for y in YEARS:
                rv = base * residual_factor(tech, y)
                mn, mx = canada_margins(y, tech)
                val = ser.get(y, base)
                rep_min, rep_max = val * mn, val * mx
                if tech in MONOTONIC_MAX_TECHS:
                    rep_max = max(rep_max, running_max)
                    running_max = rep_max
                if tech in GAS_TECHS:
                    # CER gas series dips ~2032 then recovers; a wobbling min made
                    # additions hit zero and the fleet shrink-then-regrow. Keep the
                    # Canadian gas floor monotone non-decreasing.
                    rep_min = max(rep_min, running_min_gas.get(tech, 0.0))
                    running_min_gas[tech] = rep_min
                if tech == "P_Gas_CCGT":
                    # same existing/new CCGT split as the US pools (Canada is not a
                    # no-add region, so NEW build = funnel band minus existing fleet)
                    res_rows.append(("Canada", "P_Gas_CCGT_Residual", y, round(rv, 6)))
                    res_rows.append(("Canada", "P_Gas_CCGT", y, 0.0))
                    max_rows.append(("Canada", "P_Gas_CCGT_Residual", y, round(rv, 6)))
                    new_min = max(0.0, rep_min - rv)
                    new_max = max(FORBID_EPS, rep_max - rv)
                    if new_min > 0:
                        min_rows.append(("Canada", "P_Gas_CCGT", y, round(new_min, 6)))
                    max_rows.append(("Canada", "P_Gas_CCGT", y, round(new_max, 6)))
                else:
                    res_rows.append(("Canada", tech, y, round(rv, 6)))
                    min_rows.append(("Canada", tech, y, round(rep_min, 6)))
                    max_rows.append(("Canada", tech, y, round(rep_max, 6)))

        # Canada storage: Sir Adam Beck PHS (fixed + growth-capped max) + battery
        # (CER EF2026 "Battery Storage" -> Li-Ion). Mirrors the US block 1d: PHS
        # residual = min = 2025 fleet, max grows <=PHS_MAX_GROWTH/yr; Li-Ion
        # residual decays from the 2025 battery fleet, min funnels to the CER
        # trajectory (forces the planned fleet), max left open.
        cer_bess = read_canada_storage()
        bess_2025 = cer_bess.get(2025, 0.0)
        for y in YEARS:
            res_rows.append(("Canada", "D_PHS", y, round(CANADA_PHS_GW_2025, 6)))
            min_rows.append(("Canada", "D_PHS", y, round(CANADA_PHS_GW_2025, 6)))
            max_rows.append(("Canada", "D_PHS", y, round(max(FORBID_EPS, CANADA_PHS_GW_2025 * (1.0 + PHS_MAX_GROWTH) ** (y - 2025)), 6)))
            bess = cer_bess.get(y, bess_2025)
            mn, mx = canada_margins(y)
            res_rows.append(("Canada", "D_Battery_Li-Ion", y, round(bess_2025 * residual_factor("D_Battery_Li-Ion", y), 6)))
            min_rows.append(("Canada", "D_Battery_Li-Ion", y, round(bess * mn, 6)))
            max_rows.append(("Canada", "D_Battery_Li-Ion", y, round(max(FORBID_EPS, bess * mx), 6)))

    # 3a) Other extra regions + Canada upside variants from restool potentials.
    for region in extra_regions:
        for col, info in RESTOOL_MAP.items():
            pot_val = pot.get(region, {}).get(col, 0.0)
            rep = info["rep"]
            for variant, share in info["variants"].items():
                if region == "Canada" and variant in canada_techs:
                    continue   # CER block owns rep + rooftop for Canada
                target = pot_val * share
                for y in YEARS:
                    if region == "Canada":
                        # Canada's guardrail rep is the CER-owned _Avg, so the
                        # restool rep (_Opt) is NOT funnel-carried here - a flat
                        # full-potential max let the optimiser dump 13-17 GW into
                        # single years (real Canadian build rates: ~1-2.5 GW/yr
                        # each for wind/solar). Ramp ALL restool classes on the
                        # slow Canadian window (the slope = the annual ceiling).
                        val = min(target * canada_restool_frac(y),
                                  CANADA_RESTOOL_SLOPE_GW * max(0, y - CANADA_RESTOOL_RAMP_START + 1))
                    elif variant == rep or rep is None:
                        val = target
                    else:
                        val = target * restool_frac(y)
                    max_rows.append((region, variant, y, round(val, 6)))

    # 3b) Coal in extra regions: SK lignite stays at the SaskPower 1.53 GW
    #     flat (CER shows 1.39 incl. Boundary Dam CCS — same picture).
    #     Hardcoal (NS/NB/AB) follows the CER decline to 0 by ~2035.
    #     MaxCapacity = residual (no new coal builds).
    for region in extra_regions:
        lign = CANADA_LIGNITE_GW if region == "Canada" else 0.0
        for y in YEARS:
            hard = ca_hardcoal.get(y, 0.0) if (region == "Canada" and canada_techs) else 0.0
            res_rows.append((region, "P_Coal_Lignite", y, lign))
            max_rows.append((region, "P_Coal_Lignite", y, max(FORBID_EPS, lign)))
            res_rows.append((region, "P_Coal_Hardcoal", y, round(hard, 6)))
            # FORBID_EPS floor so Canadian hardcoal (CER -> 0 by ~2035) and any
            # coal-free extra region don't flip to unlimited via the 0->999999 rule.
            max_rows.append((region, "P_Coal_Hardcoal", y, round(max(FORBID_EPS, hard), 6)))

    all_regions = pool_regions + extra_regions

    # 3c) CAES (D_CAES): per-region deployment ceiling (CAES_MAX_GW_2040),
    #     ramped linearly 0 (2025) -> ceiling (2040), flat after. Residual 0
    #     (existing ~0; McIntosh 0.11 GW negligible). No min. This bounds the
    #     long-duration store so the optimiser cannot over-build cheap CAES and
    #     must turn to the next LDES option (e.g. Redox-Flow) for the remainder.
    for region in all_regions:
        ceil40 = CAES_MAX_GW_2040.get(region, 0.0)
        for y in YEARS:
            frac = min(1.0, max(0.0, (y - 2025) / (2040 - 2025)))
            res_rows.append((region, "D_CAES", y, 0.0))
            max_rows.append((region, "D_CAES", y, round(max(FORBID_EPS, ceil40 * frac), 6)))

    # 3d) Nuclear SMR (P_Nuclear_SMR) for every region: residual 0 (no SMR fleet
    #     in 2025), TotalAnnualMaxCapacity = full PMK SMR forecast (FORBID_EPS
    #     where 0 so the 0->999999 thermal rule does not make it unlimited),
    #     TotalAnnualMinCapacity = committed "Development" SMRs. Canada carries
    #     the Saskatchewan unit (filed in the US sheet, NERC subregion MRO-CN).
    for region in all_regions:
        smr_mx = nuclear_smr_max.get(region, {})
        smr_mn = nuclear_smr_min.get(region, {})
        for y in YEARS:
            res_rows.append((region, "P_Nuclear_SMR", y, 0.0))
            mn = round(smr_mn.get(y, 0.0), 6)
            if mn > 0:
                min_rows.append((region, "P_Nuclear_SMR", y, mn))
            max_rows.append((region, "P_Nuclear_SMR", y, round(max(FORBID_EPS, smr_mx.get(y, 0.0)), 6)))

    # 4) Zero-out unmanaged power-producing techs (P_*, CHP_*) for the NA
    #    regions + Canada. Storage (D_*, S_*) and sector-coupling techs are
    #    untouched. Techs covered by sibling scripts (offshore wind, EGS) stay.
    #    Explicit zero for P_PV_Rooftop_Residential (rooftop generation lives
    #    only on P_PV_Rooftop_Commercial in this dataset).
    all_techs = pd.read_csv(TECH_SHEET()).iloc[:, 0].astype(str).tolist()
    zero_techs = sorted(
        t for t in all_techs
        if (t.startswith("P_") or t.startswith("CHP_"))
        and t not in ALL_MANAGED_TECHS
        and t not in EXTERNAL_OWNERS
    )
    zero_rows = []
    for region in all_regions:
        for tech in zero_techs:
            # Canada: techs carried by the CER block (e.g. P_Oil, P_Biomass)
            # are managed there and must not be zeroed.
            if region == "Canada" and tech in canada_techs:
                continue
            for y in YEARS:
                # FORBID_EPS, not 0: most of these unmanaged techs are FossilPower/
                # CHP/Biomass-tagged, so a 0 max is flipped to 999999 (UNLIMITED) by
                # genesysmod_bounds. Writing 0.001 actually forbids them — otherwise
                # e.g. CHP_Coal_Lignite builds ~unlimited cheap lignite in every region.
                zero_rows.append((region, tech, y, FORBID_EPS))
    max_rows.extend(zero_rows)

    # 4b) Redox flow batteries: forbid new build before REDOX_START_YEAR, then a ramping
    #     ceiling capping annual additions at REDOX_ANNUAL_CAP GW/region (residual 0, no
    #     forced min). Stops the optimiser building tens of GW of Redox up front.
    for region in all_regions:
        for y in YEARS:
            res_rows.append((region, "D_Battery_Redox", y, 0.0))
            if y < REDOX_START_YEAR:
                max_rows.append((region, "D_Battery_Redox", y, FORBID_EPS))
            else:
                max_rows.append((region, "D_Battery_Redox", y, round((y - REDOX_START_YEAR + 1) * REDOX_ANNUAL_CAP, 6)))

    # 4c) P_SOFC: open in the US pool regions (999999 = no per-region level cap;
    #     the SOFC group addition cap governs the build pace), forbidden in the
    #     extra regions (no data-center SOFC market assumed there).
    for region in pool_regions:
        for y in YEARS:
            max_rows.append((region, "P_SOFC", y, 999999.0))
    for region in extra_regions:
        for y in YEARS:
            max_rows.append((region, "P_SOFC", y, FORBID_EPS))

    # 5) National gas post-pass (US pool regions).
    #    (a) Clamp the FORCED new-build of the CCGT/OCGT/Steam group (positive
    #        increments of max(0, min - residual), summed nationally) to
    #        GAS_GROUP_MIN_CLAMP_GW per year by scaling down that year's min
    #        increments - the funnel then stays inside the group addition cap.
    #    (b) Collect the P_Gas_Engines forced new-build per year - its group
    #        addition cap follows the file whenever the file exceeds the policy
    #        cap (2.5 GW/yr from 2027).
    _res_map = {}
    for (r, t, y, v) in res_rows:
        _res_map[(r, t, y)] = v          # last write wins (matches write() order)
    _CLAMP_TECHS = ("P_Gas_CCGT", "P_Gas_OCGT", "P_Gas_Steam")
    _min_idx = {}
    for i, (r, t, y, v) in enumerate(min_rows):
        if r in pool_regions and t in _CLAMP_TECHS + ("P_Gas_Engines",):
            _min_idx[(r, t, y)] = i
    def _needed(r, t, y):
        i = _min_idx.get((r, t, y))
        if i is None:
            return 0.0
        return max(0.0, min_rows[i][3] - _res_map.get((r, t, y), 0.0))
    # (b) engines national forced-new increments (pre-clamp; engines not clamped)
    engines_inc = {}
    for y in YEARS[1:]:
        engines_inc[y] = sum(max(0.0, _needed(r, "P_Gas_Engines", y) - _needed(r, "P_Gas_Engines", y - 1))
                             for r in pool_regions)
    # (a) clamp the rest of the group
    scaled_years = []
    needed_new = {(r, t): {YEARS[0]: _needed(r, t, YEARS[0])} for r in pool_regions for t in _CLAMP_TECHS}
    for y in YEARS[1:]:
        tot_pos = sum(max(0.0, _needed(r, t, y) - _needed(r, t, y - 1))
                      for r in pool_regions for t in _CLAMP_TECHS)
        clamp_y = max(0.0, GAS_GROUP_CAP_GW.get(y, GAS_GROUP_CAP_FROM_2035) - GAS_CLAMP_BUFFER_GW)
        sc = min(1.0, clamp_y / tot_pos) if tot_pos > 0 else 1.0
        if sc < 1.0:
            scaled_years.append((y, round(tot_pos, 1), round(sc, 3)))
        for r in pool_regions:
            for t in _CLAMP_TECHS:
                delta = _needed(r, t, y) - _needed(r, t, y - 1)
                if delta > 0:
                    delta *= sc
                prev = needed_new[(r, t)][y - 1]
                needed_new[(r, t)][y] = min(max(0.0, prev + delta), _needed(r, t, y))
    for (r, t), path in needed_new.items():
        for y, nv in path.items():
            i = _min_idx.get((r, t, y))
            if i is None:
                continue
            new_min = _res_map.get((r, t, y), 0.0) + nv
            if new_min < min_rows[i][3] - 1e-9:
                min_rows[i] = (r, t, y, round(new_min, 6))
    if scaled_years:
        print("gas min clamp (forced-new > cap-%.0f): %s" % (GAS_CLAMP_BUFFER_GW,
              ", ".join("%d: %.1f GW x%.3f" % t for t in scaled_years)))

    # 6) Group annual-addition caps (base data; sensitivities override via their
    #    scenario subfolders). GasPlants excludes P_Gas_Engines (own subset).
    def _gas_group_cap(y):
        return GAS_GROUP_CAP_GW.get(y, GAS_GROUP_CAP_FROM_2035)
    def _sofc_group_cap(y):
        if y <= 2029:
            return 0.0
        return round(min(2.0, 2.0 * (y - 2029) / 6.0), 2)
    group_newcaps = [
        ("GasPlants", "USA", {y: _gas_group_cap(y) for y in YEARS},
         "gas annual-additions cap: turbine supply 6/9/10/35/35 GW 2026-2030, ramp to 60 by 2035 (TU Berlin assumption)"),
        # +0.01 GW headroom on file-driven values: the forced-new sum and the cap
        # are rounded independently, and an exactly-equal pair is one rounding
        # step away from infeasible (2026 was infeasible by 4e-4 without this).
        ("GasEngines", "USA", {y: round(max(2.5 if y >= 2027 else 0.0,
                                            engines_inc.get(y, 0.0) + (0.01 if engines_inc.get(y, 0.0) > 0 else 0.0)), 3)
                               for y in YEARS},
         "engines cap: 2.5 GW/yr from 2027, raised to the capacities-file forced additions where higher"),
        ("SOFC", "USA", {y: _sofc_group_cap(y) for y in YEARS},
         "SOFC FTM additions: 0 until 2029, ramp to 2 GW/yr by 2035 (TU Berlin assumption)"),
        ("OffshoreWind", "Canada", {y: 2.0 for y in YEARS},
         "Canadian offshore build pace cap 2 GW/yr (TU Berlin assumption)"),
    ]

    all_written_techs = ALL_MANAGED_TECHS | set(zero_techs) | REDOX_TECHS

    def write(param, rows, src, techs=None):
        path = PARAM(param)
        d = pd.read_csv(path)
        d = d.rename(columns={"Unnamed: 4": ""})
        add = pd.DataFrame([{"Region": r, "Technology": t, "Year": y, "Value": v, "": "",
                             "Unit": "GW", "Source": src, "Updated at": DATE, "Updated by": WHO}
                            for (r, t, y, v) in rows])
        add = add[d.columns]
        if SCENARIO_SUBDIR:
            # scenario mode: the base CSV is untouched; the generated rows go to
            # Par_X/<subdir>/ and the conversion upserts them over the base.
            outdir = os.path.join(os.path.dirname(path), SCENARIO_SUBDIR)
            if apply:
                os.makedirs(outdir, exist_ok=True)
                add.to_csv(os.path.join(outdir, os.path.basename(path)), index=False, lineterminator="\n")
            return 0, len(rows)
        drop = d["Region"].isin(all_regions) & d["Technology"].isin(all_written_techs if techs is None else techs)
        nd = int(drop.sum())
        d = d[~drop]
        out = pd.concat([d, add], ignore_index=True)
        if apply:
            out.to_csv(path, index=False)
        return nd, len(rows)

    def write_subset_row(param, tech, subset):
        path = os.path.join(DATA_REPO, "Data", "Parameters", "00_Sets&Tags", param + ".csv")
        d = pd.read_csv(path)
        already = ((d["Technology"] == tech) & (d["Subset"] == subset)).any()
        if already:
            return 0, 0
        d = d.rename(columns={"Unnamed: 3": ""})
        new = pd.DataFrame([{"Technology": tech, "Subset": subset, "Value": 1, "": "",
                             "Unit": "Binary", "Source": "not relevant",
                             "Updated at": DATE, "Updated by": WHO}])
        new = new[d.columns]
        out = pd.concat([d, new], ignore_index=True)
        if apply:
            out.to_csv(path, index=False)
        return 0, 1

    def write_group_min(year_to_gw, tech_subset, region_subset, src):
        path = PARAM("Par_GroupTotalAnnualMinCapacity")
        d = pd.read_csv(path)
        d = d.rename(columns={"Unnamed: 4": ""})
        drop = (d["TechnologySubset"] == tech_subset) & (d["RegionSubset"] == region_subset)
        nd = int(drop.sum())
        d = d[~drop]
        rows = [{"TechnologySubset": tech_subset, "RegionSubset": region_subset,
                 "Year": y, "Value": v, "": "", "Unit": "GW", "Source": src,
                 "Updated at": DATE, "Updated by": WHO}
                for y, v in sorted(year_to_gw.items())]
        if rows:                      # empty year_to_gw => purge-only
            out = pd.concat([d, pd.DataFrame(rows)[d.columns]], ignore_index=True)
        else:
            out = d
        if apply:
            out.to_csv(path, index=False)
        return nd, len(rows)

    def write_group_newcaps(entries):
        # base mode only: sensitivities carry their own Par_GroupTotalAnnualMaxNewCap
        # rows in their scenario subfolders (row-upsert over these at conversion)
        if SCENARIO_SUBDIR:
            return 0, 0
        path = PARAM("Par_GroupTotalAnnualMaxNewCap")
        d = pd.read_csv(path)
        d = d.rename(columns={c: "" for c in d.columns if str(c).startswith("Unnamed")})
        nd = 0
        rows = []
        for tech_subset, region_subset, year_to_gw, src in entries:
            drop = (d["TechnologySubset"] == tech_subset) & (d["RegionSubset"] == region_subset)
            nd += int(drop.sum())
            d = d[~drop]
            rows += [{"TechnologySubset": tech_subset, "RegionSubset": region_subset,
                      "Year": y, "Value": v, "": "", "Unit": "GW", "Source": src,
                      "Updated at": DATE, "Updated by": WHO}
                     for y, v in sorted(year_to_gw.items())]
        out = pd.concat([d, pd.DataFrame(rows)[d.columns]], ignore_index=True)
        if apply:
            out.to_csv(path, index=False, lineterminator="\n")
        return nd, len(rows)

    if MAX_BOOST:
        # boost applied after the persistence floor: scaling up never violates it;
        # forbid rows (<= FORBID_EPS) stay forbidden (boosting one would lift it
        # past the model's DEAD_CAP_EPS and silently re-enable the tech)
        max_rows = [(r, t, y, round(v * MAX_BOOST, 6))
                    if t.startswith(MAX_BOOST_PREFIXES) and v > FORBID_EPS
                    and (MAX_BOOST_REGIONS is None or r in MAX_BOOST_REGIONS)
                    else (r, t, y, v) for (r, t, y, v) in max_rows]

    if SCENARIO_SUBDIR:
        # Upsert semantics only OVERRIDE matching rows: a min row the base has
        # but this sensitivity does not would leak through. Emit explicit zeros
        # for every managed (region, tech, year) the sensitivity did not set.
        have = {(r, t, y) for (r, t, y, v) in min_rows}
        for r in all_regions:
            for t in sorted(all_written_techs):
                for y in YEARS:
                    if (r, t, y) not in have:
                        min_rows.append((r, t, y, 0.0))
    r1 = write("Par_ResidualCapacity", res_rows,
               "US Pools 2025 base x retirement; Nuclear=2025 LSR (PMK); Coal=US Coal Trajectory LOW; "
               "P_Gas_Steam incl. coal->gas conversions (LOW)")
    r2 = write("Par_TotalAnnualMinCapacity", min_rows,
               "US Pools widening band (min); Nuclear=LSR trajectory (PMK); Coal=US Coal Trajectory CENTRAL")
    r3 = write("Par_TotalAnnualMaxCapacity", max_rows,
               "Guardrail (2026-2035) + restool potential (2036-2040); Nuclear=LSR x headroom; "
               "Coal=US Coal Trajectory HIGH")

    r4 = write_subset_row("Par_TagTechnologyToSubsets", "P_Nuclear", "Nuclear")
    for _t, _ss in (("P_Gas_Engines", "GasEngines"),
                    ("P_Wind_Offshore_Deep", "OffshoreWind"),
                    ("P_Wind_Offshore_Shallow", "OffshoreWind"),
                    ("P_Wind_Offshore_Transitional", "OffshoreWind")):
        write_subset_row("Par_TagTechnologyToSubsets", _t, _ss)
    r6 = write_group_newcaps(group_newcaps)
    # Per-region P_Nuclear TotalAnnualMinCapacity now enforces the LSR trajectory,
    # so purge the superseded US-aggregate Nuclear group-min (add nothing).
    r5 = write_group_min({}, "Nuclear", "USA", "superseded by per-region P_Nuclear min")

    # Sample print: PJM P_Gas_CCGT (no restool ceiling — held flat post-2035);
    #               the PV cascade — _Opt is the rep (carries residual + funnel),
    #               _Avg / _Inf only take the overflow above Opt's share-of-potential.
    for ex_r, ex_t in (("California", "P_Gas_CCGT_Residual"), ("California", "P_Gas_CCGT"),
                       ("NewEngland", "P_Gas_CCGT_Residual"), ("NewEngland", "P_Gas_CCGT"),
                       ("PJM", "P_Gas_CCGT_Residual"), ("PJM", "P_Gas_CCGT"),
                       ("Canada", "P_Gas_CCGT_Residual"), ("Canada", "P_Gas_CCGT"),
                       ("SERC", "P_Wind_Onshore_Avg")):
        print(f"\nSample {ex_r} {ex_t} (GW):  Year  Residual    Min      Max")
        res = {y: v for (r, t, y, v) in res_rows if r == ex_r and t == ex_t}
        mn = {y: v for (r, t, y, v) in min_rows if r == ex_r and t == ex_t}
        mx = {y: v for (r, t, y, v) in max_rows if r == ex_r and t == ex_t}
        for y in YEARS:
            print(f"          {y}  {res.get(y,0):>9.3f}  {mn.get(y, float('nan')):>7.3f}  "
                  f"{mx.get(y, float('nan')):>7.3f}")

    print(f"\n{'APPLIED' if apply else 'DRY-RUN'}: Residual -{r1[0]}/+{r1[1]} ; "
          f"MinCap -{r2[0]}/+{r2[1]} ; MaxCap -{r3[0]}/+{r3[1]} ; "
          f"NuclearSubset -{r4[0]}/+{r4[1]} ; "
          f"NuclearGroupMin -{r5[0]}/+{r5[1]} "
          f"(pool_regions={len(pool_regions)}, extra={len(extra_regions)}, "
          f"managed_techs={len(ALL_MANAGED_TECHS)})")
    if not apply:
        print("\n(use --apply to write)")


if __name__ == "__main__":
    main()
