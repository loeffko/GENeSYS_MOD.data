"""
Register the P_Nuclear_SMR technology and copy every P_Nuclear *technology-
property* row to it: cost (Capital/Fixed/Variable), OperationalLife,
CapacityToActivityUnit, AvailabilityFactor, ramping, ProductionChangeCost,
fuel In/OutputActivityRatio, and the Sector/Subset/ReserveMargin tags. The
uranium InputActivityRatio is included so the power-only precompute bakes the
SMR fuel cost the same way it does for P_Nuclear.

NOT copied (region capacity / policy, set elsewhere): ResidualCapacity,
TotalAnnualMin/MaxCapacity, AnnualMin/MaxNewCapacity, NewCapacityExpansionStop,
RegionalBaseYearProduction. SMR's NA capacity (residual 0, min = firmly-
committed "Development" SMRs, max = full SMR forecast) comes from
add_capacity_bounds.py.

"for now" = SMR shares P_Nuclear's economics/efficiency verbatim (placeholder
until SMR-specific cost/lifetime data lands).

Also adds P_Nuclear_SMR (selected) to the Technology_selection sheet of both NA
set-filter files in "Conversion Script/", otherwise the CSV->Excel conversion
would drop the new technology even though it is in Sets_Technology.

NOT handled here: SMR capacity (residual/min/max) -> add_capacity_bounds.py;
building the InputData Excels -> the usual conversion notebook/scripts.

Idempotent. Dry-run by default; pass --apply to write.

Run:  python NA_inputs/add_nuclear_smr.py [--apply]   (run BEFORE add_capacity_bounds.py)
"""
import os, sys, csv, io
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_REPO = os.path.normpath(os.path.join(HERE, ".."))
SRC, NEW = "P_Nuclear", "P_Nuclear_SMR"
apply = "--apply" in sys.argv
PARAMS = os.path.join(DATA_REPO, "Data", "Parameters")
SETS = os.path.join(PARAMS, "00_Sets&Tags", "Sets_Technology.csv")
FILTER_DIR = os.path.join(DATA_REPO, "Conversion Script")
FILTER_FILES = ["Set_filter_file_NorthAmerica.xlsx",
                "Set_filter_file_NorthAmerica_allFuels.xlsx"]

# Technology-property CSVs: copy every row that names P_Nuclear -> P_Nuclear_SMR.
COPY = [
    "00_Sets&Tags/Par_TagTechnologyToSector.csv",
    "00_Sets&Tags/Par_TagTechnologyToSubsets.csv",
    "00_Sets&Tags/Par_CapacityToActivityUnit.csv",
    "00_Sets&Tags/Par_ReserveMarginTagTechnology.csv",
    "Par_OperationalLife/Par_OperationalLife.csv",
    "Par_CapitalCost/Par_CapitalCost.csv",
    "Par_FixedCost/Par_FixedCost.csv",
    "Par_VariableCost/Par_VariableCost.csv",
    "Par_InputActivityRatio/Par_InputActivityRatio.csv",
    "Par_OutputActivityRatio/Par_OutputActivityRatio.csv",
    "Par_AvailabilityFactor/Par_AvailabilityFactor.csv",
    "Par_ProductionChangeCost/Par_ProductionChangeCost.csv",
    "Par_RampingUpFactor/Par_RampingUpFactor.csv",
    "Par_RampingDownFactor/Par_RampingDownFactor.csv",
]

def _ends_with_newline(path):
    with open(path, "rb") as f:
        if f.seek(0, os.SEEK_END) == 0:
            return True
        f.seek(-1, os.SEEK_END)
        return f.read(1) in (b"\n", b"\r")

def copy_rows(rel):
    path = os.path.join(PARAMS, rel)
    if not os.path.isfile(path):
        return (0, "missing")
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    body = rows[1:] if rows else []
    if any(NEW in r for r in body):
        return (0, "exists")
    add = [[NEW if c == SRC else c for c in r] for r in body if SRC in r]
    if not add:
        return (0, "no P_Nuclear")
    if apply:
        buf = io.StringIO()
        csv.writer(buf, lineterminator="\n").writerows(add)
        with open(path, "a", encoding="utf-8", newline="") as f:
            if not _ends_with_newline(path):
                f.write("\n")
            f.write(buf.getvalue())
    return (len(add), "ok")

def add_to_sets():
    with open(SETS, newline="", encoding="utf-8-sig") as f:
        techs = [r[0] for r in csv.reader(f) if r]
    if NEW in techs:
        return (0, "exists")
    if apply:
        with open(SETS, "a", encoding="utf-8", newline="") as f:
            if not _ends_with_newline(SETS):
                f.write("\n")
            f.write(NEW + "\n")
    return (1, "ok")

def add_to_filters():
    """Add (P_Nuclear_SMR, selected) to the Technology_selection sheet of both
    NA set-filter files so the CSV->Excel conversion keeps the new tech."""
    out = []
    for fn in FILTER_FILES:
        path = os.path.join(FILTER_DIR, fn)
        if not os.path.isfile(path):
            out.append((fn, 0, "missing")); continue
        wb = openpyxl.load_workbook(path)
        ws = wb["Technology_selection"]
        techs = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        if NEW in techs:
            out.append((fn, 0, "exists")); wb.close(); continue
        if apply:
            ws.append([NEW, 1])   # 1 = selected, like P_Nuclear
            wb.save(path)
        out.append((fn, 1, "ok")); wb.close()
    return out

def main():
    print(f"{'APPLY' if apply else 'DRY-RUN'}: add {NEW} (copy of {SRC})\n")
    n, st = add_to_sets()
    print(f"  Sets_Technology: +{n} ({st})")
    total = 0
    for rel in COPY:
        n, st = copy_rows(rel)
        total += n
        print(f"  {rel:55s} +{n:4d} ({st})")
    print(f"\n  total property rows copied: {total}")
    print("  filter files (Technology_selection):")
    for fn, n, st in add_to_filters():
        print(f"    {fn:45s} +{n} ({st})")
    if not apply:
        print("\n(use --apply to write; run add_capacity_bounds.py afterwards for SMR capacity)")

if __name__ == "__main__":
    main()
